#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Transforma los 23 Excel de campo de Angel (Directorio/No participan/Publicos...)
al formato de importacion BePharma (Empresas/Contactos/Deals), con validaciones
para una migracion segura (dedupe por dominio, campos requeridos, log de alertas).
"""
import glob
import os
import re
import unicodedata
import openpyxl
from openpyxl.utils import get_column_letter

SRC_DIR = "Angel"
OUT_XLSX = "BePharma_Excel_Importacion_Angel.xlsx"
OPERATOR_EMAIL = "international@bepharma.org"
OPERATOR_ZONE = "angel"
EVENT_CODE = "BEPH-2026-09"
PIPELINE_ID = "908086638"

FILE_COUNTRY = {
    "Arabia Saudita Septiembre 2026.xlsx": "Arabia Saudita",
    "Australia Septiembre 2026.xlsx": "Australia",
    "Belice Septiembre 2026.xlsx": "Belice",
    "China Septiembre  al 28 de junio 2026.xlsx": "China",
    "Diversos_USA Septiembre 2026.xlsx": "Estados Unidos",
    "Emiratos Arabes Septiembre 2026.xlsx": "Emiratos Arabes Unidos",
    "Florida Septiembre 2026.xlsx": "Estados Unidos",
    "India Septiembre 2026.xlsx": "India",
    "Italia Septiembre 2026 enriquecido.xlsx": "Italia",
    "Jamaica al 16 JUN 2026 respaldo 16jun2026 13 48pm.xlsx": "Jamaica",
    "Japon Septiembre 2026.xlsx": "Japon",
    "Korea Septiembre 2026 enriquecido.xlsx": "Corea del Sur",
    "Nueva Zelanda Septiembre 2026.xlsx": "Nueva Zelanda",
    "Pakistan Septiembre 2026.xlsx": "Pakistan",
    "Republica Checa Septiembre 2026.xlsx": "Republica Checa",
    "Rusia Septiembre 2026.xlsx": "Rusia",
    "Singapur Septiembre 2026.xlsx": "Singapur",
    "SriLanka Septiembre 2026.xlsx": "Sri Lanka",
    "Tennessee Septiembre 2026 (2).xlsx": "Estados Unidos",
    "Texas Septiembre 2026.xlsx": "Estados Unidos",
    "Vietnam Septiembre 2026.xlsx": "Vietnam",
}

COUNTRY_REGION = {
    "arabia saudita": "medio_oriente", "emiratos arabes unidos": "medio_oriente",
    "turquia": "medio_oriente",
    "australia": "asia_pac", "china": "asia_pac", "india": "asia_pac",
    "japon": "asia_pac", "corea del sur": "asia_pac", "nueva zelanda": "asia_pac",
    "pakistan": "asia_pac", "singapur": "asia_pac", "sri lanka": "asia_pac",
    "vietnam": "asia_pac",
    "italia": "europa_occ",
    "republica checa": "europa_oriental", "rusia": "europa_oriental",
    "belice": "latam_norte", "jamaica": "latam_norte",
    "estados unidos": "norteamerica",
}


def strip_accents(s):
    if s is None:
        return ""
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def norm(s):
    if s is None:
        return ""
    s = strip_accents(str(s)).lower().strip()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("-", " ")
    return s


def region_for(country):
    return COUNTRY_REGION.get(norm(country), "")


HEADER_SYNONYMS = {
    "empresa": "empresa",
    "giro y exportacion": "giro", "giro": "giro",
    "seguimiento empresa": "seguimiento_empresa", "comentarios empresa": "seguimiento_empresa",
    "seguimiento": "seguimiento_empresa",
    "nombre": "nombre_contacto",
    "cargo": "cargo",
    "coment persona": "coment_persona", "comentario persona": "coment_persona",
    "comentarios sobre la persona": "coment_persona",
    "tel movil": "tel_movil", "telefono movil": "tel_movil", "celular": "tel_movil",
    "tel persona": "tel_movil",
    "coment llamada": "coment_llamada", "comentario llamada": "coment_llamada",
    "comentario seguimiento llamada": "coment_llamada",
    "seguimiento de la persona 2025": "coment_llamada",
    "email persona": "email_persona", "e mail persona": "email_persona",
    "e mail": "email_persona",
    "coment email": "coment_email", "comentario email": "coment_email",
    "comentario seguimiento e mail": "coment_email", "comentario seguimiento email": "coment_email",
    "tel empresa": "tel_empresa", "telefono empresa": "tel_empresa",
    "email empresa": "email_empresa", "email empresas": "email_empresa",
    "e mail empresa": "email_empresa", "e mail empresas": "email_empresa",
    "pagina web": "pagina_web",
    "comentarios historicos": "coment_historicos", "coment historicos": "coment_historicos",
    "lkd": "linkedin",
    "mensaje lkd": "mensaje_lkd",
    "coment lkd": "coment_lkd", "comentario seguimiento lkd": "coment_lkd",
    "imp y partic": "prioridad",
    "pais": "pais",
}


def build_column_roles(header_row_vals):
    normed = [norm(h) if h else "" for h in header_row_vals]
    tel_empresa_idx = next((i for i, h in enumerate(normed) if h in ("tel empresa", "telefono empresa")), None)
    email_empresa_idx = next((i for i, h in enumerate(normed) if h in ("email empresa", "email empresas", "e mail empresa", "e mail empresas")), None)

    roles = [None] * len(normed)
    for i, h in enumerate(normed):
        if tel_empresa_idx is not None and i == tel_empresa_idx + 1:
            roles[i] = "coment_tel_empresa"
            continue
        if email_empresa_idx is not None and i == email_empresa_idx + 1:
            roles[i] = "coment_email_empresa"
            continue
        roles[i] = HEADER_SYNONYMS.get(h)
    return roles


def find_header_row(ws, max_scan_rows=8, max_scan_cols=24):
    for r in range(1, max_scan_rows + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, max_scan_cols + 1)]
        normed = [norm(v) for v in vals]
        if "empresa" in normed and "nombre" in normed:
            return r, vals[:max_scan_cols]
    return None, None


def classify_sheet(sheet_name):
    n = norm(sheet_name)
    if "no particip" in n:
        return "no_participa", "No participa (hoja fuente: %s)" % sheet_name.strip()
    if "publico" in n or "no importante" in n:
        return "no_participa", "Publico / No importante (hoja fuente: %s)" % sheet_name.strip()
    return "activo", None


def extract_ref_timezone(ws):
    text = None
    for r in range(1, 4):
        v = ws.cell(row=r, column=2).value or ws.cell(row=r, column=1).value
        if v and "Zona horaria" in str(v):
            text = str(v)
            break
    if not text:
        return "", ""
    tz = ""
    ventana = ""
    m = re.search(r"Zona horaria:\s*(.*?)\s*\|", text)
    if m:
        tz = m.group(1).strip()
    m = re.search(r"Ventana pr[aá]ctica desde CDMX:\s*(.*?)\s*\|", text)
    if m:
        ventana = m.group(1).strip()
    return tz, ventana


EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@([A-Za-z0-9.\-]+\.[A-Za-z]{2,})")

FREEMAIL_DOMAINS = {
    "gmail.com", "googlemail.com", "hotmail.com", "hotmail.es", "hotmail.co.uk",
    "yahoo.com", "yahoo.com.mx", "yahoo.co.uk", "yahoo.es", "ymail.com", "rocketmail.com",
    "outlook.com", "outlook.es", "live.com", "live.com.mx", "msn.com",
    "aol.com", "icloud.com", "me.com", "mac.com",
    "protonmail.com", "proton.me", "gmx.com", "gmx.net", "mail.com",
    "yandex.com", "yandex.ru", "zoho.com", "qq.com", "163.com", "126.com", "naver.com",
    "web.de", "libero.it", "virgilio.it",
}


def extract_domain(*candidates):
    for c in candidates:
        if not c:
            continue
        m = EMAIL_RE.search(str(c))
        if m:
            d = m.group(1).lower()
            if d.startswith("www."):
                d = d[4:]
            if d in FREEMAIL_DOMAINS:
                continue
            return d
    return ""


SIGNIFICANT_WORD_MIN_LEN = 4
GENERIC_NAME_WORDS = {"pharma", "pharmaceuticals", "pharmaceutical", "group", "ltd", "limited",
                      "company", "co", "corp", "corporation", "inc", "international", "medical",
                      "medics", "medicals", "trading", "imports", "supplies", "laboratories",
                      "laboratorio", "industries", "industry"}


def similar_names(a, b):
    wa = {w for w in norm(a).split() if len(w) >= SIGNIFICANT_WORD_MIN_LEN and w not in GENERIC_NAME_WORDS}
    wb = {w for w in norm(b).split() if len(w) >= SIGNIFICANT_WORD_MIN_LEN and w not in GENERIC_NAME_WORDS}
    return bool(wa & wb)


def normalize_website_domain(url):
    if not url:
        return ""
    u = str(url).strip().lower()
    u = re.sub(r"^https?://", "", u)
    u = u.split("/")[0]
    if u.startswith("www."):
        u = u[4:]
    return u


def clean_phone(v):
    if v is None:
        return ""
    s = str(v).strip()
    s = s.lstrip("*").strip()
    return s


def guess_tipo_empresa(giro):
    g = norm(giro)
    if "cdmo" in g:
        return "cdmo"
    if "fabricante" in g:
        return "fabricante"
    if "distribui" in g:
        return "distribuidor"
    if "laboratorio" in g:
        return "laboratorio"
    if "investigaci" in g:
        return "investigacion"
    return ""


DECISOR_KEYWORDS = ["ceo", "chief", "president", "director", "vp", "vice president",
                    "founder", "managing director", "gerente general", "owner"]
INFLUENCIADOR_KEYWORDS = ["manager", "gerente", "head of", "jefe"]


def guess_cargo_decisor(jobtitle):
    j = norm(jobtitle)
    if not j:
        return ""
    if any(k in j for k in DECISOR_KEYWORDS):
        return "decisor"
    if any(k in j for k in INFLUENCIADOR_KEYWORDS):
        return "influenciador"
    return "desconocido"


def split_name(full):
    parts = str(full).strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def join_notes(*parts):
    vals = [str(p).strip() for p in parts if p not in (None, "")]
    return "\n---\n".join(vals)


PRIORITY_SCORE = {"alta": "80", "media": "50", "baja": "20"}


MESES = {
    "ene": 1, "enero": 1, "feb": 2, "febrero": 2, "mar": 3, "marzo": 3,
    "abr": 4, "abril": 4, "may": 5, "mayo": 5, "jun": 6, "junio": 6,
    "jul": 7, "julio": 7, "ago": 8, "agosto": 8, "sep": 9, "sept": 9, "septiembre": 9,
    "oct": 10, "octubre": 10, "nov": 11, "noviembre": 11, "dic": 12, "diciembre": 12,
}

DATE_LEAD_RE = re.compile(
    r"(?:\*\s*)?\d{1,2}\s*(?:/|de\s+|\s+)"
    r"(?:ene(?:ro)?|feb(?:rero)?|mar(?:zo)?|abr(?:il)?|may(?:o)?|jun(?:io)?|jul(?:io)?|"
    r"ago(?:sto)?|sep(?:t(?:iembre)?)?|oct(?:ubre)?|nov(?:iembre)?|dic(?:iembre)?)"
    r"|(?:lunes|martes|mi[ée]rcoles|jueves|viernes|s[áa]bado|domingo)\s+\d{1,2}/\d{1,2}/\d{4}"
    r"|\d{1,2}/\d{1,2}/\d{4}",
    re.IGNORECASE,
)

TIME_RE = re.compile(r"(\d{1,2})[:\s](\d{2})\s*((?:a|p)\.?\s?m\.?)?", re.IGNORECASE)


def parse_fecha_aproximada(chunk):
    """Extrae una fecha aproximada (YYYY-MM-DD HH:MM) del inicio de un fragmento
    de texto en espanol con formatos muy variados (26/MAY, 13 de enero, domingo
    28/06/2026, 22 junio 2026 18 16 pm, etc). Best-effort: si no reconoce un
    patron valido devuelve None -- nunca inventa una fecha."""
    head = chunk[:60]
    nhead = norm(head)

    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", head)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    else:
        m = re.search(r"\b(\d{1,2})\s*/\s*([a-z]{3,})", nhead)
        if not m:
            m = re.search(r"\b(\d{1,2})\s+de\s+([a-z]{3,})", nhead)
        if not m:
            m = re.search(r"\b(\d{1,2})\s+([a-z]{3,})\s+(\d{2,4})\b", nhead)
        if not m:
            return None
        day = int(m.group(1))
        mes_txt = m.group(2)[:3] if m.group(2)[:3] in MESES else m.group(2)
        month = MESES.get(mes_txt) or MESES.get(m.group(2))
        if not month:
            return None
        year = 2026
        if m.lastindex and m.lastindex >= 3:
            try:
                y = int(m.group(3))
                year = y if y > 100 else 2000 + y
            except (ValueError, IndexError):
                pass

    if not (1 <= day <= 31 and 1 <= month <= 12):
        return None

    hh, mm = None, None
    tail_search_area = head[head.find(str(day)) if str(day) in head else 0:]
    tm = TIME_RE.search(chunk[:80])
    if tm:
        try:
            hh = int(tm.group(1))
            mm = int(tm.group(2))
            ampm = (tm.group(3) or "").lower().replace(".", "").replace(" ", "")
            if ampm == "pm" and hh < 12:
                hh += 12
            if ampm == "am" and hh == 12:
                hh = 0
            if hh > 23:
                hh, mm = None, None
        except ValueError:
            hh, mm = None, None

    try:
        if hh is not None and mm is not None:
            return "%04d-%02d-%02d %02d:%02d" % (year, month, day, hh, mm)
        return "%04d-%02d-%02d" % (year, month, day)
    except ValueError:
        return None


def split_into_movements(text):
    """Divide un bloque de texto libre en movimientos individuales usando
    marcas de fecha reconocibles. Si no encuentra al menos 2 marcas, devuelve
    el texto completo como un unico movimiento (mejor no dividir mal que
    inventar cortes). Nunca descarta texto."""
    if not text:
        return []
    text = str(text).strip()
    if not text:
        return []
    matches = list(DATE_LEAD_RE.finditer(text))
    if len(matches) < 2:
        return [{"fecha": parse_fecha_aproximada(text), "texto": text}]
    movimientos = []
    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        chunk = text[start:end].strip(" \n\t-*")
        if chunk:
            movimientos.append({"fecha": parse_fecha_aproximada(chunk), "texto": chunk})
    return movimientos


def process_workbook(path, log):
    fname = os.path.basename(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    country_default = FILE_COUNTRY.get(fname)
    if country_default is None:
        if "turquia" in norm(fname):
            country_default = "Turquia"
        else:
            country_default = ""
            log.append(("ARCHIVO SIN PAIS MAPEADO", fname, "", "Revisar mapeo de pais manualmente"))

    companies = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row_idx, header_vals = find_header_row(ws)
        if header_row_idx is None:
            log.append(("HOJA SIN HEADER RECONOCIBLE", fname, sheet_name, "No se encontraron columnas Empresa/Nombre; hoja omitida"))
            continue
        roles = build_column_roles(header_vals)
        tz, ventana = extract_ref_timezone(ws)
        estado_prospeccion, motivo_descarte = classify_sheet(sheet_name)

        col_by_role = {}
        for i, role in enumerate(roles):
            if role and role not in col_by_role:
                col_by_role[role] = i + 1

        def val(row, role):
            c = col_by_role.get(role)
            if not c:
                return None
            return ws.cell(row=row, column=c).value

        current = None
        max_row = ws.max_row
        for r in range(header_row_idx + 1, max_row + 1):
            empresa_v = val(r, "empresa")
            nombre_v = val(r, "nombre_contacto")
            if empresa_v is None and nombre_v is None:
                continue
            if empresa_v is not None and str(empresa_v).strip():
                if current:
                    companies.append(current)
                pais_row = val(r, "pais") or country_default
                current = {
                    "empresa": str(empresa_v).strip(),
                    "giro": val(r, "giro"),
                    "seguimiento_empresa": val(r, "seguimiento_empresa"),
                    "prioridad": val(r, "prioridad"),
                    "tel_empresa": val(r, "tel_empresa"),
                    "coment_tel_empresa": val(r, "coment_tel_empresa"),
                    "email_empresa": val(r, "email_empresa"),
                    "coment_email_empresa": val(r, "coment_email_empresa"),
                    "pagina_web": val(r, "pagina_web"),
                    "coment_historicos": val(r, "coment_historicos"),
                    "pais": pais_row,
                    "estado_prospeccion": estado_prospeccion,
                    "motivo_descarte": motivo_descarte,
                    "sheet_source": sheet_name,
                    "file_source": fname,
                    "tz": tz,
                    "ventana": ventana,
                    "contactos": [],
                }
            if nombre_v is not None and str(nombre_v).strip():
                if current is None:
                    log.append(("CONTACTO SIN EMPRESA PADRE", fname, sheet_name,
                                 "Fila %d: contacto '%s' sin empresa previa; omitido" % (r, nombre_v)))
                    continue
                current["contactos"].append({
                    "nombre": str(nombre_v).strip(),
                    "cargo": val(r, "cargo"),
                    "coment_persona": val(r, "coment_persona"),
                    "tel_movil": val(r, "tel_movil"),
                    "coment_llamada": val(r, "coment_llamada"),
                    "email_persona": val(r, "email_persona"),
                    "coment_email": val(r, "coment_email"),
                    "linkedin": val(r, "linkedin"),
                    "mensaje_lkd": val(r, "mensaje_lkd"),
                    "coment_lkd": val(r, "coment_lkd"),
                })
        if current:
            companies.append(current)
    return companies


def main():
    log = []
    all_companies = []
    files = sorted(glob.glob(os.path.join(SRC_DIR, "*.xlsx")))
    print("Procesando %d archivos..." % len(files))
    for f in files:
        comps = process_workbook(f, log)
        print("  %s: %d empresas" % (os.path.basename(f), len(comps)))
        all_companies.extend(comps)

    print("\nTotal empresas extraidas (antes de dedupe): %d" % len(all_companies))

    domain_registry = {}
    name_registry = {}
    empresas_rows = []
    contactos_rows = []
    deals_rows = []
    movimientos_rows = []
    idx = 0

    for comp in all_companies:
        domain = extract_domain(comp.get("email_empresa"),
                                 *[c.get("email_persona") for c in comp["contactos"]])
        if not domain and comp.get("pagina_web"):
            domain = normalize_website_domain(comp["pagina_web"])

        skip = False
        if domain:
            existing = domain_registry.get(domain, [])
            exact_dup = next((n for n in existing if similar_names(comp["empresa"], n)), None)
            if exact_dup:
                log.append(("EMPRESA DUPLICADA (mismo dominio, nombre relacionado)", comp["file_source"], comp["sheet_source"],
                             "'%s' ya existia como '%s' con el mismo dominio (%s) -- omitida, revisar manualmente si son distintas" % (comp["empresa"], exact_dup, domain)))
                skip = True
            elif existing:
                log.append(("DOMINIO COMPARTIDO ENTRE EMPRESAS DISTINTAS", comp["file_source"], comp["sheet_source"],
                             "'%s' comparte el dominio '%s' con %s (nombres no relacionados) -- se importan ambas, verificar que el dominio sea el correcto para cada una" % (comp["empresa"], domain, existing)))
            domain_registry.setdefault(domain, []).append(comp["empresa"])
        else:
            key = norm(comp["empresa"])
            if key in name_registry:
                log.append(("EMPRESA DUPLICADA (mismo nombre, sin dominio)", comp["file_source"], comp["sheet_source"],
                             "'%s' ya existia como '%s' (ninguna tiene dominio) -- omitida, revisar manualmente si son distintas" % (comp["empresa"], name_registry[key])))
                skip = True
            else:
                name_registry[key] = comp["empresa"]

        if skip:
            continue

        idx += 1
        bp_id = "BP-ANG-%05d" % idx

        if not domain:
            log.append(("SIN DOMINIO", comp["file_source"], comp["sheet_source"],
                         "'%s' (%s) sin email ni pagina web utilizable (o solo correo gratuito tipo gmail/hotmail) -- completar dominio a mano antes de importar" % (comp["empresa"], bp_id)))

        priority_norm = norm(comp.get("prioridad"))
        score = PRIORITY_SCORE.get(priority_norm, "")

        tipo_empresa = guess_tipo_empresa(comp.get("giro"))

        has_company_contact_data = bool(clean_phone(comp.get("tel_empresa"))) or bool(comp.get("email_empresa"))
        has_any_contact_with_data = any(
            (c.get("tel_movil") or c.get("email_persona")) for c in comp["contactos"]
        )
        estado_enriquecimiento = "completo" if (has_company_contact_data and has_any_contact_with_data) else (
            "parcial" if (has_company_contact_data or has_any_contact_with_data) else "pendiente")

        motivo_descarte = comp.get("motivo_descarte") or ""

        empresas_rows.append({
            "bp_id_unico": bp_id,
            "name": comp["empresa"],
            "bp_dominio_normalizado": domain,
            "domain": normalize_website_domain(comp.get("pagina_web")) if comp.get("pagina_web") else "",
            "country": comp.get("pais") or "",
            "bp_region_comercial": region_for(comp.get("pais")),
            "bp_zona": OPERATOR_ZONE,
            "bp_tipo_empresa": tipo_empresa,
            "bp_estado_prospeccion": "nueva" if comp["estado_prospeccion"] == "activo" else "no_participa",
            "bp_estado_enriquecimiento": estado_enriquecimiento,
            "bp_fuente_enriquecimiento": "manual",
            "bp_clasificacion_claude": "",
            "bp_score_bepharma": score,
            "bp_estado_alerta": "",
            "bp_zona_horaria_pais": comp.get("tz") or "",
            "bp_ventana_llamada_mx": comp.get("ventana") or "",
            "phone": clean_phone(comp.get("tel_empresa")),
            "industry": "",
            "numberofemployees": "",
            "bp_motivo_descarte": motivo_descarte,
            "description": comp.get("giro") or "",
        })

        for c in comp["contactos"]:
            if not (c.get("nombre") or "").strip():
                continue
            firstname, lastname = split_name(c["nombre"])
            email = (c.get("email_persona") or "").strip() if c.get("email_persona") else ""
            invitacion = "enviada" if (c.get("coment_email") or comp.get("coment_email_empresa")) else "no_enviada"
            contactos_rows.append({
                "firstname": firstname,
                "lastname": lastname,
                "email": email,
                "phone": clean_phone(c.get("tel_movil")),
                "jobtitle": c.get("cargo") or "",
                "company": comp["empresa"],
                "associatedcompanydomain": domain,
                "bp_cargo_decisor": guess_cargo_decisor(c.get("cargo")),
                "bp_email_verificado": "en_verificacion" if email else "",
                "bp_telefono_verificado": "en_verificacion" if clean_phone(c.get("tel_movil")) else "",
                "bp_fuente_contacto": "manual",
                "bp_estado_invitacion": invitacion,
                "linkedin_url": c.get("linkedin") or "",
            })
            if not email:
                log.append(("CONTACTO SIN EMAIL", comp["file_source"], comp["sheet_source"],
                            "'%s' en '%s' sin email -- HubSpot puede rechazar el contacto si no tiene ningun identificador" % (c["nombre"], comp["empresa"])))

            for canal, texto in (
                ("Comentario contacto", c.get("coment_persona")),
                ("Llamada contacto", c.get("coment_llamada")),
                ("Email contacto", c.get("coment_email")),
                ("LinkedIn mensaje", c.get("mensaje_lkd")),
                ("LinkedIn comentario", c.get("coment_lkd")),
            ):
                for mov in split_into_movements(texto):
                    movimientos_rows.append({
                        "dealname": "%s - %s" % (comp["empresa"], EVENT_CODE),
                        "associatedcompanydomain": domain,
                        "empresa": comp["empresa"],
                        "contacto_relacionado": c["nombre"],
                        "canal": canal,
                        "hs_timestamp": mov["fecha"] or "",
                        "hs_note_body": "[%s | %s] %s" % (canal, c["nombre"], mov["texto"]),
                        "hubspot_owner_id": OPERATOR_EMAIL,
                    })
                    if not mov["fecha"]:
                        log.append(("MOVIMIENTO SIN FECHA RECONOCIDA", comp["file_source"], comp["sheet_source"],
                                     "Contacto '%s' en '%s' (%s): no se pudo detectar fecha en el texto -- completar hs_timestamp a mano o se usara la fecha de importacion" % (c["nombre"], comp["empresa"], canal)))

        dealstage_label = "Nueva empresa" if comp["estado_prospeccion"] == "activo" else "No participa"
        deals_rows.append({
            "dealname": "%s - %s" % (comp["empresa"], EVENT_CODE),
            "associatedcompanydomain": domain,
            "pipeline": PIPELINE_ID,
            "dealstage": dealstage_label,
            "hubspot_owner_id": OPERATOR_EMAIL,
            "bp_evento_codigo": EVENT_CODE,
            "bp_evento_nombre": "",
            "bp_evento_estado": "",
            "bp_zona": OPERATOR_ZONE,
            "bp_evento_fecha_inicio": "",
            "bp_evento_fecha_cierre": "",
            "bp_evento_meta_contactar": "",
            "bp_evento_meta_confirmar": "",
            "bp_evento_paises": "",
            "closedate": "",
        })

        for canal, texto in (
            ("Seguimiento general", comp.get("seguimiento_empresa")),
            ("Llamada empresa", comp.get("coment_tel_empresa")),
            ("Email empresa", comp.get("coment_email_empresa")),
            ("Historico", comp.get("coment_historicos")),
        ):
            for mov in split_into_movements(texto):
                movimientos_rows.append({
                    "dealname": "%s - %s" % (comp["empresa"], EVENT_CODE),
                    "associatedcompanydomain": domain,
                    "empresa": comp["empresa"],
                    "contacto_relacionado": "",
                    "canal": canal,
                    "hs_timestamp": mov["fecha"] or "",
                    "hs_note_body": "[%s] %s" % (canal, mov["texto"]),
                    "hubspot_owner_id": OPERATOR_EMAIL,
                })
                if not mov["fecha"]:
                    log.append(("MOVIMIENTO SIN FECHA RECONOCIDA", comp["file_source"], comp["sheet_source"],
                                 "Empresa '%s' (%s): no se pudo detectar fecha en el texto -- completar hs_timestamp a mano o se usara la fecha de importacion" % (comp["empresa"], canal)))

    print("\nEmpresas finales (post-dedupe): %d" % len(empresas_rows))
    print("Contactos: %d" % len(contactos_rows))
    print("Deals: %d" % len(deals_rows))
    print("Alertas en log: %d" % len(log))
    print("Movimientos: %d" % len(movimientos_rows))

    write_output(empresas_rows, contactos_rows, deals_rows, movimientos_rows, log)


def write_output(empresas_rows, contactos_rows, deals_rows, movimientos_rows, log):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def add_sheet(name, rows, columns):
        ws = wb.create_sheet(name)
        ws.append(columns)
        for row in rows:
            ws.append([row.get(col, "") for col in columns])
        for i, col in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(14, min(40, len(col) + 4))
        ws.freeze_panes = "A2"

    empresas_cols = ["bp_id_unico", "name", "bp_dominio_normalizado", "domain", "country",
                     "bp_region_comercial", "bp_zona", "bp_tipo_empresa", "bp_estado_prospeccion",
                     "bp_estado_enriquecimiento", "bp_fuente_enriquecimiento", "bp_clasificacion_claude",
                     "bp_score_bepharma", "bp_estado_alerta", "bp_zona_horaria_pais",
                     "bp_ventana_llamada_mx", "phone", "industry", "numberofemployees",
                     "bp_motivo_descarte", "description"]
    contactos_cols = ["firstname", "lastname", "email", "phone", "jobtitle", "company",
                      "associatedcompanydomain", "bp_cargo_decisor", "bp_email_verificado",
                      "bp_telefono_verificado", "bp_fuente_contacto", "bp_estado_invitacion",
                      "linkedin_url"]
    deals_cols = ["dealname", "associatedcompanydomain", "pipeline", "dealstage",
                  "hubspot_owner_id", "bp_evento_codigo", "bp_evento_nombre", "bp_evento_estado",
                  "bp_zona", "bp_evento_fecha_inicio", "bp_evento_fecha_cierre",
                  "bp_evento_meta_contactar", "bp_evento_meta_confirmar", "bp_evento_paises",
                  "closedate", "description"]

    movimientos_cols = ["dealname", "associatedcompanydomain", "empresa", "contacto_relacionado",
                        "canal", "hs_timestamp", "hs_note_body", "hubspot_owner_id"]

    add_sheet("EMPRESAS", empresas_rows, empresas_cols)
    add_sheet("CONTACTOS", contactos_rows, contactos_cols)
    add_sheet("DEALS", deals_rows, deals_cols)
    add_sheet("MOVIMIENTOS_DEAL", movimientos_rows, movimientos_cols)

    ws_log = wb.create_sheet("REVISAR ANTES DE IMPORTAR")
    ws_log.append(["Tipo de alerta", "Archivo", "Hoja", "Detalle"])
    for row in log:
        ws_log.append(list(row))
    for i, w in enumerate([35, 30, 25, 90], start=1):
        ws_log.column_dimensions[get_column_letter(i)].width = w
    ws_log.freeze_panes = "A2"

    wb.save(OUT_XLSX)
    print("\nArchivo guardado: %s" % OUT_XLSX)


if __name__ == "__main__":
    main()
