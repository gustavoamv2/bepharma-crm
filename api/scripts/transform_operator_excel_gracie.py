#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Transforma los Excel de campo de Gracie (Directorio/No participan/Publicos...,
mas archivos irregulares: hojas sueltas, multi-pais por archivo, LinkedIn) al
formato de importacion BePharma (Empresas/Contactos/Deals/Movimientos).

Copiado y adaptado de transform_operator_excel_carlos.py (que ya trae el fix
de columnas posicionales y el dedupe unificado). Cambios nuevos para Gracie:

- Los archivos de Gracie son MUCHO mas heterogeneos que los de Angel/Carlos:
  hojas con nombres genericos ("Hoja1"/"Hoja2"/"Hoja 1"), un archivo con una
  hoja por pais ("Caribe e Islas.xlsx"), un archivo "Colombia..." que en
  realidad tiene Colombia en Hoja1 y Uruguay en Hoja2, y un archivo
  "LinkedIn BePharma Gracie.xlsx" que mezcla MUCHOS paises en una sola hoja
  usando el patron "PRIORIDAD\nPais" dentro de la celda "Imp y partic".
- Fix: se agrega extract_country_from_row1() -- heuristica generica que lee
  el texto de la fila 1 de cada hoja (que casi siempre trae el nombre del
  pais/estado, con sufijos variables tipo "— Septiembre 2026", "2026",
  "+3 horas") y lo usa como pais por defecto de esa hoja, en vez de asumir
  un solo pais por archivo. Esto resuelve automaticamente el caso
  Colombia/Uruguay sin necesitar un mapeo a mano.
- Fix: se agrega deteccion de "PRIORIDAD\nPais" en la columna Imp y partic
  (visto en LinkedIn BePharma Gracie.xlsx) como override de pais POR FILA,
  con prioridad sobre el pais de la hoja/archivo.
- Fix: varios estados de EE.UU. vienen como archivos propios (Alabama,
  California, Georgia, Illinois, Massachusetts, North Carolina, Nueva York,
  Pennsylvania) -- se normalizan a country="Estados Unidos" (para la
  propiedad country de HubSpot) conservando el estado como tag al inicio
  del giro/descripcion, y a bp_region_comercial="norteamerica".
- Fix: nuevos sinonimos de columnas vistos en los archivos de Gracie
  (Argentina, Panama, Colombia) + regla generica de fallback: cualquier
  header con "seguimiento" que no matchee un sinonimo exacto se clasifica
  como seguimiento_empresa (si menciona "empresa") o coment_llamada (si no).
- Fix: nuevo campo coment_roberto (columna "Comentarios Roberto", vista en
  Panama) -- se preserva por separado en vez de pisar coment_historicos.
"""
import glob
import os
import re
import unicodedata
import openpyxl
from openpyxl.utils import get_column_letter

SRC_DIR = "Gracie"
OUT_XLSX = "BePharma_Excel_Importacion_Gracie.xlsx"
OPERATOR_EMAIL = "worldwide@bepharma.org"
OPERATOR_ZONE = "gracie"
OPERATOR_PREFIX = "BP-GRA"
EVENT_CODE = "BEPH-2026-09"
PIPELINE_ID = "908086638"

# Fallback SOLO si extract_country_from_row1() no logra leer nada util para
# esa hoja/archivo (deberia ser raro -- casi todos los archivos de Gracie
# traen el pais/estado en la fila 1).
FILE_COUNTRY_FALLBACK = {
    "Alabama Septiembre 2026.xlsx": "Alabama",
    "Argentina - Yes.xlsx": "Argentina",
    "Belgica Septiembre 2026.xlsx": "Belgica",
    "Bulgaria Septiembre 2026.xlsx": "Bulgaria",
    "California Septiembre 2026 2.xlsx": "California",
    "Croacia Septiembre 2026.xlsx": "Croacia",
    "Dinamarca Septiembre 2026.xlsx": "Dinamarca",
    "Egipto Septiembre 2026.xlsx": "Egipto",
    "Escocia Septiembre 2026.xlsx": "Escocia",
    "Finlandia Septiembre 2026.xlsx": "Finlandia",
    "Georgia Septiembre 2026.xlsx": "Georgia",  # estado de EE.UU., no el pais
    "Haiti Septiembre 2026.xlsx": "Haiti",
    "Illinois Septiembre 2026.xlsx": "Illinois",
    "Inglaterra Septiembre 2026.xlsx": "Inglaterra",
    "Irlanda Septiembre 2026.xlsx": "Irlanda",
    "Massachusetts Septiembre 2026.xlsx": "Massachusetts",
    "North Carolina Septiembre 2026.xlsx": "North Carolina",
    "Noruega Septiembre 2026.xlsx": "Noruega",
    "Nueva York Septiembre 2026  final.xlsx": "Nueva York",
    "Panama.xlsx": "Panama",
    "Pennsylvania Septiembre 2026.xlsx": "Pennsylvania",
    "Rumania Septiembre 2026.xlsx": "Rumania",
    "Taiwan Septiembre 2026.xlsx": "Taiwan",
    "Caribe e Islas.xlsx": "",  # cada hoja trae su propio pais (ver sheet_name)
    "LinkedIn BePharma Gracie.xlsx": "",  # multi-pais, se resuelve por fila
    "Colombia Gracie septiembre 2026.xlsx": "",  # Hoja1=Colombia, Hoja2=Uruguay via row1
}

# Estados de EE.UU. que llegaron como archivo/hoja propia -- se normaliza la
# propiedad "country" a Estados Unidos, conservando el estado como tag.
STATE_TO_COUNTRY = {
    "alabama": "Estados Unidos",
    "california": "Estados Unidos",
    "georgia": "Estados Unidos",
    "illinois": "Estados Unidos",
    "massachusetts": "Estados Unidos",
    "north carolina": "Estados Unidos",
    "new york": "Estados Unidos",
    "nueva york": "Estados Unidos",
    "pennsylvania": "Estados Unidos",
}

COUNTRY_REGION = {
    "alabama": "norteamerica", "california": "norteamerica", "georgia": "norteamerica",
    "illinois": "norteamerica", "massachusetts": "norteamerica", "north carolina": "norteamerica",
    "new york": "norteamerica", "nueva york": "norteamerica", "pennsylvania": "norteamerica",
    "estados unidos": "norteamerica", "ee uu": "norteamerica", "ee.uu.": "norteamerica", "eeuu": "norteamerica",
    "belgica": "europa_occ", "finlandia": "europa_occ", "inglaterra": "europa_occ",
    "irlanda": "europa_occ", "noruega": "europa_occ", "escocia": "europa_occ",
    "bulgaria": "europa_oriental", "croacia": "europa_oriental", "rumania": "europa_oriental",
    "dinamarca": "europa_occ",
    "egipto": "africa",
    "taiwan": "asia_pacifico",
    "argentina": "latam_sur", "colombia": "latam_norte", "uruguay": "latam_sur",
    "panama": "latam_norte", "haiti": "latam_caribe",
    "trinidad y tobago": "latam_caribe", "barbados": "latam_caribe", "bahamas": "latam_caribe",
    "surinam": "latam_caribe", "antigua": "latam_caribe",
    "antigua . aruba . curacao": "latam_caribe", "antigua aruba curacao": "latam_caribe",
}


def strip_accents(s):
    if s is None:
        return ""
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def norm(s):
    if s is None:
        return ""
    s = strip_accents(str(s)).lower().strip()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("-", " ")
    s = s.replace("·", " ").replace(",", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def region_for(country):
    return COUNTRY_REGION.get(norm(country), "")


def build_file_country_map(files):
    mapping = {}
    for f in files:
        mapping[os.path.basename(f)] = FILE_COUNTRY_FALLBACK.get(os.path.basename(f), "")
    return mapping


def _clean_row1_segment(s):
    s = s.strip()
    # prefijo "Directorio " (ej. "Directorio Egipto 2026")
    s = re.sub(r"(?i)^directorio\s+", "", s).strip()
    # sufijo "/ EE.UU." (ej. "New York / EE.UU.")
    s = re.sub(r"(?i)\s*/\s*ee\.?\s*uu\.?\s*$", "", s).strip()
    # sufijo "+N horas" (ej. "Uruguay +3 horas")
    s = re.sub(r"(?i)\s*\+\s*\d+\s*horas?\s*$", "", s).strip()
    # sufijo/segmento "(Directorio) Septiembre/Sep NNNN"
    s = re.sub(r"(?i)\s*(directorio\s*)?(septiembre|sep\.?)\s*\d{4}\s*$", "", s).strip()
    # segmento entero "No participan" / "Publicos y no importantes"
    if re.match(r"(?i)^no\s*participan$", s) or re.match(r"(?i)^p[uú]blicos?\s*y\s*no\s*importantes?$", s):
        return ""
    # año suelto al final (ej. "Croacia 2026" -> "Croacia")
    s = re.sub(r"\s*\d{4}\s*$", "", s).strip()
    return s


def extract_country_from_row1(ws):
    # Heuristica generica: casi todas las hojas de Gracie traen el nombre
    # del pais/estado en la celda A1, mezclado de forma variable con
    # "Directorio"/"No participan"/"Publicos y no importantes"/fechas/husos
    # horarios, a veces antes y a veces despues (separado por em-dash,
    # guion largo, "--" o "/"). Se parte el texto en segmentos, se limpia
    # cada uno de los tokens conocidos, y se usa el primer segmento no vacio
    # que quede. Si no queda nada usable, devuelve "" (el llamador decide el
    # fallback: pais de una hoja anterior del mismo archivo, o el default).
    v = ws.cell(row=1, column=1).value
    if not v:
        return ""
    text = str(v).strip()
    if not text or len(text) > 80:
        return ""
    segments = re.split(r"\s*(?:—|--)\s*|\s+-\s+", text)
    kept = []
    for seg in segments:
        cleaned = _clean_row1_segment(seg)
        if cleaned:
            kept.append(cleaned)
    if not kept:
        return ""
    result = kept[0]
    if len(result) > 40:
        return ""
    if result.isupper():
        result = result.title()
    return result


# Prioridad + pais combinados en una sola celda, visto en
# "LinkedIn BePharma Gracie.xlsx" columna Imp y partic, ej: "ALTA\nCroacia"
def split_priority_country(raw):
    # OJO: la segunda linea de "Imp y partic" NO siempre es un pais -- en
    # varios archivos (ej. Belgica) se usa el mismo patron "ALTA\n..." para
    # guardar codigos de eventos pasados (ej. "VP2018 / VP2019 / BP2021 /
    # BPM2026"), no un pais. Solo en el archivo LinkedIn BePharma Gracie.xlsx
    # la segunda linea es realmente un pais (ej. "ALTA\nCroacia"). Se
    # distingue por contenido: un pais real es texto sin digitos y sin "/"
    # como separador de multiples codigos; si la segunda linea tiene digitos
    # o barras, se descarta como pais y se ignora (se deja que el pais de la
    # hoja/archivo decida).
    if raw is None:
        return None, None
    text = str(raw).strip()
    if "\n" not in text:
        return text, None
    parts = [p.strip() for p in text.split("\n") if p.strip()]
    if len(parts) < 2:
        return (parts[0] if parts else None), None
    candidate = parts[1]
    looks_like_place = bool(re.match(r"^[A-Za-zÀ-ÿ .]+$", candidate)) and len(candidate) <= 30
    country = candidate if looks_like_place else None
    return parts[0], country


HEADER_SYNONYMS = {
    "empresa": "empresa",
    "giro y exportacion": "giro", "giro": "giro",
    "seguimiento empresa": "seguimiento_empresa", "comentarios empresa": "seguimiento_empresa",
    "seguimiento": "seguimiento_empresa", "seguimiento de la empresa 2026": "seguimiento_empresa",
    "seguimiento empresa 2026": "seguimiento_empresa",
    "comentarios sobre la empresa": "seguimiento_empresa", "comentarios de la empresa": "seguimiento_empresa",
    "nombre": "nombre_contacto",
    "cargo": "cargo", "carrgo": "cargo",
    "coment persona": "coment_persona", "comentario persona": "coment_persona",
    "comentarios sobre la persona": "coment_persona",
    "tel movil": "tel_movil", "telefono movil": "tel_movil", "celular": "tel_movil",
    "tel persona": "tel_movil", "telefono persona": "tel_movil",
    "coment llamada": "coment_llamada", "comentario llamada": "coment_llamada",
    "comentario seguimiento llamada": "coment_llamada",
    "seguimiento de la persona 2025": "coment_llamada",
    "email persona": "email_persona", "e mail persona": "email_persona",
    "e mail": "email_persona",
    "coment email": "coment_email", "comentario email": "coment_email",
    "comentario seguimiento e mail": "coment_email", "comentario seguimiento email": "coment_email",
    "tel empresa": "tel_empresa", "telefono empresa": "tel_empresa",
    "email empresa": "email_empresa", "email empresas": "email_empresa",
    "e mail empresa": "email_empresa", "e mail empresas": "email_empresa",
    "mail de la empresa": "email_empresa",
    "pagina web": "pagina_web",
    "comentarios historicos": "coment_historicos", "coment historicos": "coment_historicos",
    "comentarios roberto": "coment_roberto",
    "lkd": "linkedin",
    "mensaje lkd": "mensaje_lkd",
    "coment lkd": "coment_lkd", "comentario seguimiento lkd": "coment_lkd",
    "imp y partic": "prioridad",
    "pais": "pais",
}


def build_column_roles(header_row_vals):
    normed = [norm(h) if h else "" for h in header_row_vals]
    tel_empresa_idx = next((i for i, h in enumerate(normed) if h in ("tel empresa", "telefono empresa")), None)
    email_empresa_idx = next((i for i, h in enumerate(normed) if h in ("email empresa", "email empresas", "e mail empresa", "e mail empresas", "mail de la empresa")), None)

    roles = [None] * len(normed)
    seen_before = set()
    for i, h in enumerate(normed):
        is_positional_target = (tel_empresa_idx is not None and i == tel_empresa_idx + 1) or \
                                (email_empresa_idx is not None and i == email_empresa_idx + 1)
        is_blank_or_dup = (h == "") or (h in seen_before)
        if is_positional_target and is_blank_or_dup:
            roles[i] = "coment_tel_empresa" if (tel_empresa_idx is not None and i == tel_empresa_idx + 1) else "coment_email_empresa"
        elif h in HEADER_SYNONYMS:
            roles[i] = HEADER_SYNONYMS[h]
        elif "seguimiento" in h:
            # Fallback generico: variantes no previstas tipo "Seguimiento 2025
            # de Gracie" / "Seguimiento de Grace 2026" -- si menciona
            # "empresa" es seguimiento de empresa, si no, de la persona.
            roles[i] = "seguimiento_empresa" if "empresa" in h else "coment_llamada"
        else:
            roles[i] = None
        if h:
            seen_before.add(h)
    return roles


def find_header_row(ws, max_scan_rows=8, max_scan_cols=24):
    for r in range(1, max_scan_rows + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, max_scan_cols + 1)]
        normed = [norm(v) for v in vals]
        if "empresa" in normed and "nombre" in normed:
            return r, vals[:max_scan_cols]
    return None, None


def classify_sheet(sheet_name):
    n = norm(sheet_name)
    if "no particip" in n:
        return "no_participa", "No participa (hoja fuente: %s)" % sheet_name.strip()
    if "publico" in n or "no importante" in n:
        return "no_participa", "Publico / No importante (hoja fuente: %s)" % sheet_name.strip()
    return "activo", None


def extract_ref_timezone(ws):
    text = None
    for r in range(1, 4):
        v = ws.cell(row=r, column=2).value or ws.cell(row=r, column=1).value
        if v and "Zona horaria" in str(v):
            text = str(v)
            break
    if not text:
        return "", ""
    tz = ""
    ventana = ""
    m = re.search(r"Zona horaria:\s*(.*?)\s*\|", text)
    if m:
        tz = m.group(1).strip()
    m = re.search(r"Ventana pr[aá]ctica desde CDMX:\s*(.*?)\s*\|", text)
    if m:
        ventana = m.group(1).strip()
    return tz, ventana


EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@([A-Za-z0-9.\-]+\.[A-Za-z]{2,})")

FREEMAIL_DOMAINS = {
    "gmail.com", "googlemail.com", "hotmail.com", "hotmail.es", "hotmail.co.uk",
    "yahoo.com", "yahoo.com.mx", "yahoo.co.uk", "yahoo.es", "ymail.com", "rocketmail.com",
    "outlook.com", "outlook.es", "live.com", "live.com.mx", "msn.com",
    "aol.com", "icloud.com", "me.com", "mac.com",
    "protonmail.com", "proton.me", "gmx.com", "gmx.net", "mail.com",
    "yandex.com", "yandex.ru", "zoho.com", "qq.com", "163.com", "126.com", "naver.com",
    "web.de", "libero.it", "virgilio.it",
}


def extract_domain(*candidates):
    for c in candidates:
        if not c:
            continue
        m = EMAIL_RE.search(str(c))
        if m:
            d = m.group(1).lower()
            if d.startswith("www."):
                d = d[4:]
            if d in FREEMAIL_DOMAINS:
                continue
            return d
    return ""


SIGNIFICANT_WORD_MIN_LEN = 4
GENERIC_NAME_WORDS = {"pharma", "pharmaceuticals", "pharmaceutical", "group", "ltd", "limited",
                      "company", "co", "corp", "corporation", "inc", "international", "medical",
                      "medics", "medicals", "trading", "imports", "supplies", "laboratories",
                      "laboratorio", "industries", "industry", "gmbh", "sa", "sarl", "ag", "srl"}


def similar_names(a, b):
    wa = {w for w in norm(a).split() if len(w) >= SIGNIFICANT_WORD_MIN_LEN and w not in GENERIC_NAME_WORDS}
    wb = {w for w in norm(b).split() if len(w) >= SIGNIFICANT_WORD_MIN_LEN and w not in GENERIC_NAME_WORDS}
    return bool(wa & wb)


def normalize_website_domain(url):
    if not url:
        return ""
    u = str(url).strip().lower()
    u = re.sub(r"^https?://", "", u)
    u = u.split("/")[0]
    if u.startswith("www."):
        u = u[4:]
    return u


def clean_phone(v):
    if v is None:
        return ""
    s = str(v).strip()
    s = s.lstrip("*").strip()
    return s


def guess_tipo_empresa(giro):
    g = norm(giro)
    if "cdmo" in g:
        return "cdmo"
    if "fabricante" in g:
        return "fabricante"
    if "distribui" in g:
        return "distribuidor"
    if "laboratorio" in g:
        return "laboratorio"
    if "investigaci" in g:
        return "investigacion"
    return ""


DECISOR_KEYWORDS = ["ceo", "chief", "president", "director", "vp", "vice president",
                    "founder", "managing director", "gerente general", "owner", "geschaftsfuhrer"]
INFLUENCIADOR_KEYWORDS = ["manager", "gerente", "head of", "jefe"]


def guess_cargo_decisor(jobtitle):
    j = norm(jobtitle)
    if not j:
        return ""
    if any(k in j for k in DECISOR_KEYWORDS):
        return "decisor"
    if any(k in j for k in INFLUENCIADOR_KEYWORDS):
        return "influenciador"
    return "desconocido"


def split_name(full):
    parts = str(full).strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def join_notes(*parts):
    vals = [str(p).strip() for p in parts if p not in (None, "")]
    return "\n---\n".join(vals)


PRIORITY_SCORE = {"alta": "80", "media": "50", "baja": "20"}

MESES = {
    "ene": 1, "enero": 1, "feb": 2, "febrero": 2, "mar": 3, "marzo": 3,
    "abr": 4, "abril": 4, "may": 5, "mayo": 5, "jun": 6, "junio": 6,
    "jul": 7, "julio": 7, "ago": 8, "agosto": 8, "sep": 9, "sept": 9, "septiembre": 9,
    "oct": 10, "octubre": 10, "nov": 11, "noviembre": 11, "dic": 12, "diciembre": 12,
}

DATE_LEAD_RE = re.compile(
    r"(?:\*\s*)?\d{1,2}\s*(?:/|de\s+|\s+)"
    r"(?:ene(?:ro)?|feb(?:rero)?|mar(?:zo)?|abr(?:il)?|may(?:o)?|jun(?:io)?|jul(?:io)?|"
    r"ago(?:sto)?|sep(?:t(?:iembre)?)?|oct(?:ubre)?|nov(?:iembre)?|dic(?:iembre)?)"
    r"|(?:lunes|martes|mi[ée]rcoles|jueves|viernes|s[áa]bado|domingo)\s+\d{1,2}/\d{1,2}/\d{4}"
    r"|\d{1,2}/\d{1,2}/\d{4}",
    re.IGNORECASE,
)

TIME_RE = re.compile(r"(\d{1,2})[:\s](\d{2})\s*((?:a|p)\.?\s?m\.?)?", re.IGNORECASE)


def parse_fecha_aproximada(chunk):
    head = chunk[:60]
    nhead = norm(head)

    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", head)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    else:
        m = re.search(r"\b(\d{1,2})\s*/\s*([a-z]{3,})", nhead)
        if not m:
            m = re.search(r"\b(\d{1,2})\s+de\s+([a-z]{3,})", nhead)
        if not m:
            m = re.search(r"\b(\d{1,2})\s+([a-z]{3,})\s+(\d{2,4})\b", nhead)
        if not m:
            return None
        day = int(m.group(1))
        mes_txt = m.group(2)[:3] if m.group(2)[:3] in MESES else m.group(2)
        month = MESES.get(mes_txt) or MESES.get(m.group(2))
        if not month:
            return None
        year = 2026
        if m.lastindex and m.lastindex >= 3:
            try:
                y = int(m.group(3))
                year = y if y > 100 else 2000 + y
            except (ValueError, IndexError):
                pass

    if not (1 <= day <= 31 and 1 <= month <= 12):
        return None

    hh, mm = None, None
    tm = TIME_RE.search(chunk[:80])
    if tm:
        try:
            hh = int(tm.group(1))
            mm = int(tm.group(2))
            ampm = (tm.group(3) or "").lower().replace(".", "").replace(" ", "")
            if ampm == "pm" and hh < 12:
                hh += 12
            if ampm == "am" and hh == 12:
                hh = 0
            if hh > 23:
                hh, mm = None, None
        except ValueError:
            hh, mm = None, None

    try:
        if hh is not None and mm is not None:
            return "%04d-%02d-%02d %02d:%02d" % (year, month, day, hh, mm)
        return "%04d-%02d-%02d" % (year, month, day)
    except ValueError:
        return None


def split_into_movements(text):
    if not text:
        return []
    text = str(text).strip()
    if not text:
        return []
    matches = list(DATE_LEAD_RE.finditer(text))
    if len(matches) < 2:
        return [{"fecha": parse_fecha_aproximada(text), "texto": text}]
    movimientos = []
    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        chunk = text[start:end].strip(" \n\t-*")
        if chunk:
            movimientos.append({"fecha": parse_fecha_aproximada(chunk), "texto": chunk})
    return movimientos


def process_workbook(path, file_country_map, log):
    fname = os.path.basename(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    file_default_country = file_country_map.get(fname, "")
    # Se recuerda el ultimo pais detectado con exito EN ESTE ARCHIVO -- las
    # hojas "No participan"/"Publicos..." muchas veces no repiten el nombre
    # del pais en su fila 1 (ej. row1 = "No participan" a secas); en ese caso
    # se hereda el pais de una hoja anterior del mismo archivo (tipicamente
    # Directorio, que openpyxl procesa primero) en vez de perderlo.
    last_good_country = file_default_country

    companies = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row_idx, header_vals = find_header_row(ws)
        if header_row_idx is None:
            log.append(("HOJA SIN HEADER RECONOCIBLE", fname, sheet_name, "No se encontraron columnas Empresa/Nombre; hoja omitida"))
            continue

        extracted = extract_country_from_row1(ws)
        if extracted:
            sheet_country = extracted
            last_good_country = extracted
        else:
            sheet_country = last_good_country
        if not sheet_country:
            log.append(("HOJA SIN PAIS DETECTABLE", fname, sheet_name, "No se pudo determinar el pais/estado de esta hoja automaticamente; revisar y completar bp_region_comercial/country a mano"))

        roles = build_column_roles(header_vals)
        tz, ventana = extract_ref_timezone(ws)
        estado_prospeccion, motivo_descarte = classify_sheet(sheet_name)

        col_by_role = {}
        for i, role in enumerate(roles):
            if role and role not in col_by_role:
                col_by_role[role] = i + 1

        def val(row, role):
            c = col_by_role.get(role)
            if not c:
                return None
            return ws.cell(row=row, column=c).value

        current = None
        max_row = ws.max_row
        for r in range(header_row_idx + 1, max_row + 1):
            empresa_v = val(r, "empresa")
            nombre_v = val(r, "nombre_contacto")
            if empresa_v is None and nombre_v is None:
                continue
            if empresa_v is not None and str(empresa_v).strip():
                if current:
                    companies.append(current)
                prioridad_raw = val(r, "prioridad")
                prio_code, prio_country = split_priority_country(prioridad_raw)
                pais_row = prio_country or sheet_country
                current = {
                    "empresa": str(empresa_v).strip(),
                    "giro": val(r, "giro"),
                    "seguimiento_empresa": val(r, "seguimiento_empresa"),
                    "prioridad": prio_code,
                    "tel_empresa": val(r, "tel_empresa"),
                    "coment_tel_empresa": val(r, "coment_tel_empresa"),
                    "email_empresa": val(r, "email_empresa"),
                    "coment_email_empresa": val(r, "coment_email_empresa"),
                    "pagina_web": val(r, "pagina_web"),
                    "coment_historicos": val(r, "coment_historicos"),
                    "coment_roberto": val(r, "coment_roberto"),
                    "pais": pais_row,
                    "estado_prospeccion": estado_prospeccion,
                    "motivo_descarte": motivo_descarte,
                    "sheet_source": sheet_name,
                    "file_source": fname,
                    "tz": tz,
                    "ventana": ventana,
                    "contactos": [],
                }
            if nombre_v is not None and str(nombre_v).strip():
                if current is None:
                    log.append(("CONTACTO SIN EMPRESA PADRE", fname, sheet_name,
                                 "Fila %d: contacto '%s' sin empresa previa; omitido" % (r, nombre_v)))
                    continue
                current["contactos"].append({
                    "nombre": str(nombre_v).strip(),
                    "cargo": val(r, "cargo"),
                    "coment_persona": val(r, "coment_persona"),
                    "tel_movil": val(r, "tel_movil"),
                    "coment_llamada": val(r, "coment_llamada"),
                    "email_persona": val(r, "email_persona"),
                    "coment_email": val(r, "coment_email"),
                    "linkedin": val(r, "linkedin"),
                    "mensaje_lkd": val(r, "mensaje_lkd"),
                    "coment_lkd": val(r, "coment_lkd"),
                })
        if current:
            companies.append(current)
    return companies


def resolve_country_and_region(pais_raw):
    # Devuelve (country_para_hubspot, region, tag_para_giro)
    key = norm(pais_raw)
    if key in STATE_TO_COUNTRY:
        return STATE_TO_COUNTRY[key], region_for(pais_raw), pais_raw
    return pais_raw, region_for(pais_raw), ""


def main():
    log = []
    all_companies = []
    files = sorted(glob.glob(os.path.join(SRC_DIR, "*.xlsx")))
    file_country_map = build_file_country_map(files)
    print("Procesando %d archivos..." % len(files))
    for f in files:
        comps = process_workbook(f, file_country_map, log)
        print("  %s: %d empresas" % (os.path.basename(f), len(comps)))
        all_companies.extend(comps)

    print("\nTotal empresas extraidas (antes de dedupe): %d" % len(all_companies))

    domain_registry = {}
    name_registry = {}
    empresas_rows = []
    contactos_rows = []
    deals_rows = []
    movimientos_rows = []
    idx = 0

    for comp in all_companies:
        domain = extract_domain(comp.get("email_empresa"),
                                 *[c.get("email_persona") for c in comp["contactos"]])
        if not domain and comp.get("pagina_web"):
            domain = normalize_website_domain(comp["pagina_web"])

        skip = False
        name_key = norm(comp["empresa"])

        if domain:
            existing_by_domain = domain_registry.get(domain, [])
            similar = next((n for n in existing_by_domain if similar_names(comp["empresa"], n)), None)
            if similar:
                log.append(("EMPRESA DUPLICADA (mismo dominio, nombre relacionado)", comp["file_source"], comp["sheet_source"],
                             "'%s' ya existia como '%s' con el mismo dominio (%s) -- omitida" % (comp["empresa"], similar, domain)))
                skip = True
            elif existing_by_domain:
                log.append(("DOMINIO COMPARTIDO ENTRE EMPRESAS DISTINTAS", comp["file_source"], comp["sheet_source"],
                             "'%s' comparte el dominio '%s' con %s (nombres no relacionados) -- se importan ambas, verificar" % (comp["empresa"], domain, existing_by_domain)))

        if not skip:
            existing_by_name = name_registry.get(name_key, [])
            conflict_free = next((e for e in existing_by_name if (not domain or not e["domain"] or domain == e["domain"])), None)
            if conflict_free:
                log.append(("EMPRESA DUPLICADA (mismo nombre)", comp["file_source"], comp["sheet_source"],
                             "'%s' ya existia (dominio previo: '%s', este: '%s') -- omitida" % (comp["empresa"], conflict_free["domain"], domain)))
                skip = True
            else:
                if existing_by_name:
                    log.append(("NOMBRE IDENTICO, DOMINIOS DISTINTOS", comp["file_source"], comp["sheet_source"],
                                 "'%s' tiene el mismo nombre que otra empresa ya vista pero con dominio distinto -- se importan ambas, verificar" % comp["empresa"]))
                name_registry.setdefault(name_key, []).append({"domain": domain, "name": comp["empresa"]})
                if domain:
                    domain_registry.setdefault(domain, []).append(comp["empresa"])

        if skip:
            continue

        idx += 1
        bp_id = "%s-%05d" % (OPERATOR_PREFIX, idx)

        if not domain:
            log.append(("SIN DOMINIO", comp["file_source"], comp["sheet_source"],
                         "'%s' (%s) sin email ni pagina web utilizable (o solo correo gratuito tipo gmail/hotmail) -- completar dominio a mano antes de importar" % (comp["empresa"], bp_id)))

        priority_norm = norm(comp.get("prioridad"))
        score = PRIORITY_SCORE.get(priority_norm, "")
        tipo_empresa = guess_tipo_empresa(comp.get("giro"))

        country_hubspot, region, state_tag = resolve_country_and_region(comp.get("pais"))
        giro_final = comp.get("giro") or ""
        if state_tag:
            giro_final = "[%s] %s" % (state_tag, giro_final)

        has_company_contact_data = bool(clean_phone(comp.get("tel_empresa"))) or bool(comp.get("email_empresa"))
        has_any_contact_with_data = any(
            (c.get("tel_movil") or c.get("email_persona")) for c in comp["contactos"]
        )
        estado_enriquecimiento = "completo" if (has_company_contact_data and has_any_contact_with_data) else (
            "parcial" if (has_company_contact_data or has_any_contact_with_data) else "pendiente")

        motivo_descarte = comp.get("motivo_descarte") or ""

        empresas_rows.append({
            "bp_id_unico": bp_id,
            "name": comp["empresa"],
            "bp_dominio_normalizado": domain,
            "domain": normalize_website_domain(comp.get("pagina_web")) if comp.get("pagina_web") else "",
            "country": country_hubspot or "",
            "bp_region_comercial": region,
            "bp_zona": OPERATOR_ZONE,
            "bp_tipo_empresa": tipo_empresa,
            "bp_estado_prospeccion": "nueva" if comp["estado_prospeccion"] == "activo" else "no_participa",
            "bp_estado_enriquecimiento": estado_enriquecimiento,
            "bp_fuente_enriquecimiento": "manual",
            "bp_clasificacion_claude": "",
            "bp_score_bepharma": score,
            "bp_estado_alerta": "",
            "bp_zona_horaria_pais": comp.get("tz") or "",
            "bp_ventana_llamada_mx": comp.get("ventana") or "",
            "phone": clean_phone(comp.get("tel_empresa")),
            "industry": "",
            "numberofemployees": "",
            "bp_motivo_descarte": motivo_descarte,
            "description": giro_final,
        })

        for c in comp["contactos"]:
            if not (c.get("nombre") or "").strip():
                continue
            firstname, lastname = split_name(c["nombre"])
            email = (c.get("email_persona") or "").strip() if c.get("email_persona") else ""
            invitacion = "enviada" if (c.get("coment_email") or comp.get("coment_email_empresa")) else "no_enviada"
            contactos_rows.append({
                "firstname": firstname,
                "lastname": lastname,
                "email": email,
                "phone": clean_phone(c.get("tel_movil")),
                "jobtitle": c.get("cargo") or "",
                "company": comp["empresa"],
                "associatedcompanydomain": domain,
                "bp_cargo_decisor": guess_cargo_decisor(c.get("cargo")),
                "bp_email_verificado": "en_verificacion" if email else "",
                "bp_telefono_verificado": "en_verificacion" if clean_phone(c.get("tel_movil")) else "",
                "bp_fuente_contacto": "manual",
                "bp_estado_invitacion": invitacion,
                "linkedin_url": c.get("linkedin") or "",
            })
            if not email:
                log.append(("CONTACTO SIN EMAIL", comp["file_source"], comp["sheet_source"],
                            "'%s' en '%s' sin email -- HubSpot puede rechazar el contacto si no tiene ningun identificador" % (c["nombre"], comp["empresa"])))

            for canal, texto in (
                ("Comentario contacto", c.get("coment_persona")),
                ("Llamada contacto", c.get("coment_llamada")),
                ("Email contacto", c.get("coment_email")),
                ("LinkedIn mensaje", c.get("mensaje_lkd")),
                ("LinkedIn comentario", c.get("coment_lkd")),
            ):
                for mov in split_into_movements(texto):
                    movimientos_rows.append({
                        "dealname": "%s - %s" % (comp["empresa"], EVENT_CODE),
                        "associatedcompanydomain": domain,
                        "empresa": comp["empresa"],
                        "contacto_relacionado": c["nombre"],
                        "canal": canal,
                        "hs_timestamp": mov["fecha"] or "",
                        "hs_note_body": "[%s | %s] %s" % (canal, c["nombre"], mov["texto"]),
                        "hubspot_owner_id": OPERATOR_EMAIL,
                    })
                    if not mov["fecha"]:
                        log.append(("MOVIMIENTO SIN FECHA RECONOCIDA", comp["file_source"], comp["sheet_source"],
                                     "Contacto '%s' en '%s' (%s): no se pudo detectar fecha -- completar a mano" % (c["nombre"], comp["empresa"], canal)))

        dealstage_label = "Nueva empresa" if comp["estado_prospeccion"] == "activo" else "No participa"
        deals_rows.append({
            "dealname": "%s - %s" % (comp["empresa"], EVENT_CODE),
            "associatedcompanydomain": domain,
            "pipeline": PIPELINE_ID,
            "dealstage": dealstage_label,
            "hubspot_owner_id": OPERATOR_EMAIL,
            "bp_evento_codigo": EVENT_CODE,
            "bp_evento_nombre": "",
            "bp_evento_estado": "",
            "bp_zona": OPERATOR_ZONE,
            "bp_evento_fecha_inicio": "",
            "bp_evento_fecha_cierre": "",
            "bp_evento_meta_contactar": "",
            "bp_evento_meta_confirmar": "",
            "bp_evento_paises": "",
            "closedate": "",
        })

        for canal, texto in (
            ("Seguimiento general", comp.get("seguimiento_empresa")),
            ("Llamada empresa", comp.get("coment_tel_empresa")),
            ("Email empresa", comp.get("coment_email_empresa")),
            ("Historico", comp.get("coment_historicos")),
            ("Notas Roberto", comp.get("coment_roberto")),
        ):
            for mov in split_into_movements(texto):
                movimientos_rows.append({
                    "dealname": "%s - %s" % (comp["empresa"], EVENT_CODE),
                    "associatedcompanydomain": domain,
                    "empresa": comp["empresa"],
                    "contacto_relacionado": "",
                    "canal": canal,
                    "hs_timestamp": mov["fecha"] or "",
                    "hs_note_body": "[%s] %s" % (canal, mov["texto"]),
                    "hubspot_owner_id": OPERATOR_EMAIL,
                })
                if not mov["fecha"]:
                    log.append(("MOVIMIENTO SIN FECHA RECONOCIDA", comp["file_source"], comp["sheet_source"],
                                 "Empresa '%s' (%s): no se pudo detectar fecha -- completar a mano" % (comp["empresa"], canal)))

    print("\nEmpresas finales (post-dedupe): %d" % len(empresas_rows))
    print("Contactos: %d" % len(contactos_rows))
    print("Deals: %d" % len(deals_rows))
    print("Alertas en log: %d" % len(log))
    print("Movimientos: %d" % len(movimientos_rows))

    write_output(empresas_rows, contactos_rows, deals_rows, movimientos_rows, log)


def write_output(empresas_rows, contactos_rows, deals_rows, movimientos_rows, log):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def add_sheet(name, rows, columns):
        ws = wb.create_sheet(name)
        ws.append(columns)
        for row in rows:
            ws.append([row.get(col, "") for col in columns])
        for i, col in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(14, min(40, len(col) + 4))
        ws.freeze_panes = "A2"

    empresas_cols = ["bp_id_unico", "name", "bp_dominio_normalizado", "domain", "country",
                     "bp_region_comercial", "bp_zona", "bp_tipo_empresa", "bp_estado_prospeccion",
                     "bp_estado_enriquecimiento", "bp_fuente_enriquecimiento", "bp_clasificacion_claude",
                     "bp_score_bepharma", "bp_estado_alerta", "bp_zona_horaria_pais",
                     "bp_ventana_llamada_mx", "phone", "industry", "numberofemployees",
                     "bp_motivo_descarte", "description"]
    contactos_cols = ["firstname", "lastname", "email", "phone", "jobtitle", "company",
                      "associatedcompanydomain", "bp_cargo_decisor", "bp_email_verificado",
                      "bp_telefono_verificado", "bp_fuente_contacto", "bp_estado_invitacion",
                      "linkedin_url"]
    deals_cols = ["dealname", "associatedcompanydomain", "pipeline", "dealstage",
                  "hubspot_owner_id", "bp_evento_codigo", "bp_evento_nombre", "bp_evento_estado",
                  "bp_zona", "bp_evento_fecha_inicio", "bp_evento_fecha_cierre",
                  "bp_evento_meta_contactar", "bp_evento_meta_confirmar", "bp_evento_paises", "closedate"]
    movimientos_cols = ["dealname", "associatedcompanydomain", "empresa", "contacto_relacionado",
                        "canal", "hs_timestamp", "hs_note_body", "hubspot_owner_id"]

    add_sheet("EMPRESAS", empresas_rows, empresas_cols)
    add_sheet("CONTACTOS", contactos_rows, contactos_cols)
    add_sheet("DEALS", deals_rows, deals_cols)
    add_sheet("MOVIMIENTOS_DEAL", movimientos_rows, movimientos_cols)

    ws_log = wb.create_sheet("REVISAR ANTES DE IMPORTAR")
    ws_log.append(["Tipo de alerta", "Archivo", "Hoja", "Detalle"])
    for row in log:
        ws_log.append(list(row))
    for i, w in enumerate([35, 30, 25, 90], start=1):
        ws_log.column_dimensions[get_column_letter(i)].width = w
    ws_log.freeze_panes = "A2"

    wb.save(OUT_XLSX)
    print("\nArchivo guardado: %s" % OUT_XLSX)


if __name__ == "__main__":
    main()
