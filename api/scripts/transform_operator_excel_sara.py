#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Transforma los Excel de campo de Sara al formato de importacion BePharma
(Empresas/Contactos/Deals/Movimientos).

Copiado y adaptado de transform_operator_excel_gracie.py (que ya trae
extract_country_from_row1, dedupe unificado, columnas posicionales). Cambios
nuevos para Sara:

- Fix: separador "·" (punto medio) agregado a la deteccion de segmentos de
  la fila 1 (ej. "MALASIA 2026 · NO PARTICIPAN", "Venezuela · Septiembre
  2026 · No participan") -- antes solo se separaba por em-dash/guion.
- Fix: frase "distribuidoras adicionales" agregada como token de "ruido" a
  limpiar en la fila 1 (visto en Mexico Distribuidoras...xlsx, donde las
  hojas "No participan"/"Publicos" pierden el nombre del pais "Mexico" y
  solo queda "Distribuidoras Adicionales" en su fila 1) -- sin este fix se
  hubiera tomado "Distribuidoras Adicionales" como si fuera el pais.
- Fix: "No interesados" agregado como sinonimo de hoja de no-participacion
  en classify_sheet() (visto en Mexico al 9 junio 2026.xlsx).
- Fix: MULTI-CONTACTO EN UNA SOLA CELDA -- varios archivos de Sara (ej.
  Colombia, Mexico) meten 2+ nombres de contacto separados por saltos de
  linea en la MISMA celda "Nombre" (ej. "Adriana Panche\n\n\nBoris Torres"),
  con los cargos igual de separados por saltos de linea en la celda "Cargo"
  correspondiente. Si no se separan, HubSpot recibiria un contacto con un
  nombre ilegible tipo "Adriana Panche Boris Torres". Fix: se detecta esto y
  se generan contactos individuales; el email/telefono/comentarios de la
  fila solo se asignan al PRIMER contacto detectado (no se puede saber con
  certeza a cual de los nombres corresponde ese dato compartido) -- los
  demas quedan sin esos datos y se marcan via la alerta ya existente
  "CONTACTO SIN EMAIL" para revision manual.
- Nuevos sinonimos de columnas: "comentarios sobre persona" (sin "la"),
  "movil" (solo, visto en Colombia/Mexico No interesados).
"""
import glob
import os
import re
import unicodedata
import openpyxl
from openpyxl.utils import get_column_letter

SRC_DIR = "Sara"
OUT_XLSX = "BePharma_Excel_Importacion_Sara.xlsx"
OPERATOR_EMAIL = "global@bepharma.org"
OPERATOR_ZONE = "sara"
OPERATOR_PREFIX = "BP-SAR"
EVENT_CODE = "BEPH-2026-09"
PIPELINE_ID = "908086638"

# Fallback SOLO si extract_country_from_row1() no logra leer nada util --
# casi todos los archivos de Sara traen el pais/estado en la fila 1.
FILE_COUNTRY_FALLBACK = {
    "Colombia 3 junio 2026.xlsx": "Colombia",
    "Colorado Septiembre 2026.xlsx": "Colorado",
    "Grecia Septiembre 2026.xlsx": "Grecia",
    "Guyana Septiembre 2026.xlsx": "Guyana",
    "Hungria Septiembre 2026.xlsx": "Hungria",
    "Indiana Septiembre 2026.xlsx": "Indiana",
    "Indonesia 2 junio 2026.xlsx": "Indonesia",
    "Kansas Septiembre 2026.xlsx": "Kansas",
    "Malasia Septiembre 2026.xlsx": "Malasia",
    "Mexico Distribuidoras Septiembre 2026.xlsx": "Mexico",
    "Mexico al 9 junio 2026.xlsx": "Mexico",
    "New Jersey Septiembre 2026.xlsx": "New Jersey",
    "Paises Bajos Septiembre 2026.xlsx": "Paises Bajos",
    "Polonia Septiembre 2026.xlsx": "Polonia",
    "PuertoRico Septiembre 2026.xlsx": "Puerto Rico",
    "Suecia 10 jun 2026.xlsx": "Suecia",
    "Tailandia Septiembre 2026.xlsx": "Tailandia",
    "Venezuela Septiembre 2026.xlsx": "Venezuela",
    # "Perг 1 junio 2026.xlsx" tiene el nombre de archivo con un caracter mal
    # codificado (deberia decir "Peru"); no hace falta mapearlo a mano porque
    # su fila 1 ("Peru 2026") se lee bien via extract_country_from_row1().
}

STATE_TO_COUNTRY = {
    "colorado": "Estados Unidos",
    "indiana": "Estados Unidos",
    "kansas": "Estados Unidos",
    "new jersey": "Estados Unidos",
}

COUNTRY_REGION = {
    "colorado": "norteamerica", "indiana": "norteamerica", "kansas": "norteamerica",
    "new jersey": "norteamerica", "estados unidos": "norteamerica",
    "ee uu": "norteamerica", "ee.uu.": "norteamerica", "eeuu": "norteamerica",
    "grecia": "europa_occ", "paises bajos": "europa_occ", "suecia": "europa_occ",
    "hungria": "europa_oriental", "polonia": "europa_oriental",
    "indonesia": "asia_pacifico", "malasia": "asia_pacifico", "tailandia": "asia_pacifico",
    "colombia": "latam_norte", "mexico": "latam_norte", "venezuela": "latam_norte",
    "peru": "latam_sur",
    "guyana": "latam_caribe", "puerto rico": "latam_caribe",
}


def strip_accents(s):
    if s is None:
        return ""
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def norm(s):
    if s is None:
        return ""
    s = strip_accents(str(s)).lower().strip()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("-", " ")
    s = s.replace("·", " ").replace(",", " ").replace(".", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def region_for(country):
    return COUNTRY_REGION.get(norm(country), "")


def build_file_country_map(files):
    mapping = {}
    for f in files:
        mapping[os.path.basename(f)] = FILE_COUNTRY_FALLBACK.get(os.path.basename(f), "")
    return mapping


JUNK_SEGMENT_PHRASES = ("distribuidoras adicionales",)


def _clean_row1_segment(s):
    s = s.strip()
    s = re.sub(r"(?i)^directorio\s+", "", s).strip()
    s = re.sub(r"(?i)\s*/\s*ee\.?\s*uu\.?\s*$", "", s).strip()
    s = re.sub(r"(?i)\s*\+\s*\d+\s*horas?\s*$", "", s).strip()
    s = re.sub(r"(?i)\s*(directorio\s*)?(septiembre|sep\.?)\s*\d{4}\s*$", "", s).strip()
    if re.match(r"(?i)^no\s*participan$", s) or re.match(r"(?i)^p[uú]blicos?\s*y\s*no\s*importantes?$", s):
        return ""
    if re.match(r"(?i)^no\s*interesados?$", s):
        return ""
    if norm(s) in JUNK_SEGMENT_PHRASES:
        return ""
    if norm(s) == "bepharma":
        return ""
    s = re.sub(r"\s*\d{4}\s*$", "", s).strip()
    return s


def extract_country_from_row1(ws):
    v = ws.cell(row=1, column=1).value
    if not v:
        return ""
    text = str(v).strip()
    if not text or len(text) > 80:
        return ""
    segments = re.split(r"\s*(?:—|--|·)\s*|\s+-\s+", text)
    kept = []
    for seg in segments:
        cleaned = _clean_row1_segment(seg)
        if cleaned:
            kept.append(cleaned)
    if not kept:
        return ""
    result = kept[0]
    if len(result) > 40:
        return ""
    if result.isupper():
        result = result.title()
    return result


def split_priority_country(raw):
    # La celda "Imp y partic" puede traer la prioridad y un segundo dato
    # pegados con salto de linea, en CUALQUIER orden segun el archivo:
    # "ALTA\nCroacia" (LinkedIn Gracie, prioridad primero) o
    # "VP2018 / VP2019 / BPM2026\nALTA" (Paises Bajos, prioridad al final).
    # Se busca la parte que sea literalmente alta/media/baja para saber cual
    # es la prioridad real; el resto solo se toma como pais si parece un
    # nombre de lugar (sin digitos ni "/", no es "alta/media/baja").
    if raw is None:
        return None, None
    text = str(raw).strip()
    if "\n" not in text:
        return text, None
    parts = [p.strip() for p in text.split("\n") if p.strip()]
    if len(parts) < 2:
        return (parts[0] if parts else None), None
    prio_idx = next((i for i, p in enumerate(parts) if norm(p) in ("alta", "media", "baja")), 0)
    prio_code = parts[prio_idx]
    country = None
    for i, cand in enumerate(parts):
        if i == prio_idx:
            continue
        if norm(cand) in ("alta", "media", "baja"):
            continue
        looks_like_place = bool(re.match(r"^[A-Za-zÀ-ÿ .]+$", cand)) and len(cand) <= 30
        if looks_like_place:
            country = cand
            break
    return prio_code, country


HEADER_SYNONYMS = {
    "empresa": "empresa",
    "giro y exportacion": "giro", "giro": "giro",
    "seguimiento empresa": "seguimiento_empresa", "comentarios empresa": "seguimiento_empresa",
    "seguimiento": "seguimiento_empresa", "seguimiento de la empresa 2026": "seguimiento_empresa",
    "seguimiento empresa 2026": "seguimiento_empresa",
    "comentarios sobre la empresa": "seguimiento_empresa", "comentarios de la empresa": "seguimiento_empresa",
    "nombre": "nombre_contacto", "nombre persona": "nombre_contacto",
    "cargo": "cargo", "carrgo": "cargo",
    "coment persona": "coment_persona", "comentario persona": "coment_persona",
    "comentarios sobre la persona": "coment_persona", "comentarios sobre persona": "coment_persona",
    "comentario de persona": "coment_persona",
    "tel movil": "tel_movil", "telefono movil": "tel_movil", "celular": "tel_movil",
    "tel persona": "tel_movil", "telefono persona": "tel_movil", "movil": "tel_movil",
    "coment llamada": "coment_llamada", "comentario llamada": "coment_llamada",
    "comentario seguimiento llamada": "coment_llamada",
    "seguimiento de la persona 2025": "coment_llamada",
    "email persona": "email_persona", "e mail persona": "email_persona",
    "e mail": "email_persona", "correo personal": "email_persona",
    "coment email": "coment_email", "comentario email": "coment_email",
    "comentario seguimiento e mail": "coment_email", "comentario seguimiento email": "coment_email",
    "tel empresa": "tel_empresa", "telefono empresa": "tel_empresa", "tel oficina": "tel_empresa",
    "email empresa": "email_empresa", "email empresas": "email_empresa",
    "e mail empresa": "email_empresa", "e mail empresas": "email_empresa",
    "mail de la empresa": "email_empresa", "correo empresa": "email_empresa",
    "pagina web": "pagina_web",
    "comentarios historicos": "coment_historicos", "coment historicos": "coment_historicos",
    "comentarios roberto": "coment_roberto",
    "lkd": "linkedin",
    "mensaje lkd": "mensaje_lkd",
    "coment lkd": "coment_lkd", "comentario seguimiento lkd": "coment_lkd",
    "imp y partic": "prioridad",
    "pais": "pais",
}


def build_column_roles(header_row_vals):
    normed = [norm(h) if h else "" for h in header_row_vals]
    tel_empresa_idx = next((i for i, h in enumerate(normed) if h in ("tel empresa", "telefono empresa")), None)
    email_empresa_idx = next((i for i, h in enumerate(normed) if h in ("email empresa", "email empresas", "e mail empresa", "e mail empresas", "mail de la empresa")), None)

    roles = [None] * len(normed)
    seen_before = set()
    for i, h in enumerate(normed):
        is_positional_target = (tel_empresa_idx is not None and i == tel_empresa_idx + 1) or \
                                (email_empresa_idx is not None and i == email_empresa_idx + 1)
        is_blank_or_dup = (h == "") or (h in seen_before)
        if is_positional_target and is_blank_or_dup:
            roles[i] = "coment_tel_empresa" if (tel_empresa_idx is not None and i == tel_empresa_idx + 1) else "coment_email_empresa"
        elif h in HEADER_SYNONYMS:
            roles[i] = HEADER_SYNONYMS[h]
        elif "seguimiento" in h:
            roles[i] = "seguimiento_empresa" if "empresa" in h else "coment_llamada"
        else:
            roles[i] = None
        if h:
            seen_before.add(h)
    return roles


def find_header_row(ws, max_scan_rows=8, max_scan_cols=24):
    for r in range(1, max_scan_rows + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, max_scan_cols + 1)]
        normed = [norm(v) for v in vals]
        has_empresa = any("empresa" in h.split() for h in normed)
        has_nombre = any("nombre" in h.split() for h in normed)
        if has_empresa and has_nombre:
            return r, vals[:max_scan_cols]
    return None, None


def classify_sheet(sheet_name):
    n = norm(sheet_name)
    if "no particip" in n or "no interesad" in n:
        return "no_participa", "No participa (hoja fuente: %s)" % sheet_name.strip()
    if "publico" in n or "no importante" in n:
        return "no_participa", "Publico / No importante (hoja fuente: %s)" % sheet_name.strip()
    return "activo", None


def extract_ref_timezone(ws):
    text = None
    for r in range(1, 4):
        v = ws.cell(row=r, column=2).value or ws.cell(row=r, column=1).value
        if v and "Zona horaria" in str(v):
            text = str(v)
            break
    if not text:
        return "", ""
    tz = ""
    ventana = ""
    m = re.search(r"Zona horaria:\s*(.*?)\s*\|", text)
    if m:
        tz = m.group(1).strip()
    m = re.search(r"Ventana pr[aá]ctica desde CDMX:\s*(.*?)\s*\|", text)
    if m:
        ventana = m.group(1).strip()
    return tz, ventana


EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@([A-Za-z0-9.\-]+\.[A-Za-z]{2,})")

FREEMAIL_DOMAINS = {
    "gmail.com", "googlemail.com", "hotmail.com", "hotmail.es", "hotmail.co.uk",
    "yahoo.com", "yahoo.com.mx", "yahoo.co.uk", "yahoo.es", "ymail.com", "rocketmail.com",
    "outlook.com", "outlook.es", "live.com", "live.com.mx", "msn.com",
    "aol.com", "icloud.com", "me.com", "mac.com",
    "protonmail.com", "proton.me", "gmx.com", "gmx.net", "mail.com",
    "yandex.com", "yandex.ru", "zoho.com", "qq.com", "163.com", "126.com", "naver.com",
    "web.de", "libero.it", "virgilio.it",
}


def extract_domain(*candidates):
    for c in candidates:
        if not c:
            continue
        m = EMAIL_RE.search(str(c))
        if m:
            d = m.group(1).lower()
            if d.startswith("www."):
                d = d[4:]
            if d in FREEMAIL_DOMAINS:
                continue
            return d
    return ""


SIGNIFICANT_WORD_MIN_LEN = 4
GENERIC_NAME_WORDS = {"pharma", "pharmaceuticals", "pharmaceutical", "group", "ltd", "limited",
                      "company", "co", "corp", "corporation", "inc", "international", "medical",
                      "medics", "medicals", "trading", "imports", "supplies", "laboratories",
                      "laboratorio", "industries", "industry", "gmbh", "sa", "sarl", "ag", "srl"}


def similar_names(a, b):
    wa = {w for w in norm(a).split() if len(w) >= SIGNIFICANT_WORD_MIN_LEN and w not in GENERIC_NAME_WORDS}
    wb = {w for w in norm(b).split() if len(w) >= SIGNIFICANT_WORD_MIN_LEN and w not in GENERIC_NAME_WORDS}
    return bool(wa & wb)


def normalize_website_domain(url):
    if not url:
        return ""
    u = str(url).strip().lower()
    u = re.sub(r"^https?://", "", u)
    u = u.split("/")[0]
    if u.startswith("www."):
        u = u[4:]
    # Validacion basica: si no parece un dominio real (sin punto, con
    # espacios, ej. texto tipo "Sin Pagina"/"No tiene" metido en la celda
    # de pagina web en vez de dejarla vacia), se descarta en vez de guardar
    # basura como si fuera un dominio.
    if " " in u or "." not in u or not re.match(r"^[a-z0-9.\-]+$", u):
        return ""
    return u


def clean_phone(v):
    if v is None:
        return ""
    s = str(v).strip()
    s = s.lstrip("*").strip()
    return s


def guess_tipo_empresa(giro):
    g = norm(giro)
    if "cdmo" in g:
        return "cdmo"
    if "fabricante" in g:
        return "fabricante"
    if "distribui" in g:
        return "distribuidor"
    if "laboratorio" in g:
        return "laboratorio"
    if "investigaci" in g:
        return "investigacion"
    return ""


DECISOR_KEYWORDS = ["ceo", "chief", "president", "director", "vp", "vice president",
                    "founder", "managing director", "gerente general", "owner", "geschaftsfuhrer"]
INFLUENCIADOR_KEYWORDS = ["manager", "gerente", "head of", "jefe"]


def guess_cargo_decisor(jobtitle):
    j = norm(jobtitle)
    if not j:
        return ""
    if any(k in j for k in DECISOR_KEYWORDS):
        return "decisor"
    if any(k in j for k in INFLUENCIADOR_KEYWORDS):
        return "influenciador"
    return "desconocido"


def split_name(full):
    parts = str(full).strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def split_multi_names(nombre_raw, cargo_raw):
    # Sara mete a veces 2+ contactos en la MISMA celda "Nombre" separados por
    # saltos de linea (con "Cargo" separado igual, mismo orden). Devuelve una
    # lista de (nombre, cargo) -- el resto de los datos de la fila (email,
    # tel, comentarios) solo se le asignan al primero en el llamador, porque
    # no hay forma confiable de saber a cual de los nombres corresponden.
    names = [n.strip() for n in str(nombre_raw).split("\n") if n.strip()] if nombre_raw else []
    if len(names) <= 1:
        return [(str(nombre_raw).strip() if nombre_raw else "", cargo_raw)]
    cargos = [c.strip() for c in str(cargo_raw).split("\n") if c.strip()] if cargo_raw else []
    pairs = []
    for i, name in enumerate(names):
        cargo = cargos[i] if i < len(cargos) else ""
        pairs.append((name, cargo))
    return pairs


def join_notes(*parts):
    vals = [str(p).strip() for p in parts if p not in (None, "")]
    return "\n---\n".join(vals)


PRIORITY_SCORE = {"alta": "80", "media": "50", "baja": "20"}

MESES = {
    "ene": 1, "enero": 1, "feb": 2, "febrero": 2, "mar": 3, "marzo": 3,
    "abr": 4, "abril": 4, "may": 5, "mayo": 5, "jun": 6, "junio": 6,
    "jul": 7, "julio": 7, "ago": 8, "agosto": 8, "sep": 9, "sept": 9, "septiembre": 9,
    "oct": 10, "octubre": 10, "nov": 11, "noviembre": 11, "dic": 12, "diciembre": 12,
}

DATE_LEAD_RE = re.compile(
    r"(?:\*\s*)?\d{1,2}\s*(?:/|de\s+|\s+)"
    r"(?:ene(?:ro)?|feb(?:rero)?|mar(?:zo)?|abr(?:il)?|may(?:o)?|jun(?:io)?|jul(?:io)?|"
    r"ago(?:sto)?|sep(?:t(?:iembre)?)?|oct(?:ubre)?|nov(?:iembre)?|dic(?:iembre)?)"
    r"|(?:lunes|martes|mi[ée]rcoles|jueves|viernes|s[áa]bado|domingo)\s+\d{1,2}/\d{1,2}/\d{4}"
    r"|\d{1,2}/\d{1,2}/\d{4}",
    re.IGNORECASE,
)

TIME_RE = re.compile(r"(\d{1,2})[:\s](\d{2})\s*((?:a|p)\.?\s?m\.?)?", re.IGNORECASE)


def parse_fecha_aproximada(chunk):
    head = chunk[:60]
    nhead = norm(head)

    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", head)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    else:
        m = re.search(r"\b(\d{1,2})\s*/\s*([a-z]{3,})", nhead)
        if not m:
            m = re.search(r"\b(\d{1,2})\s+de\s+([a-z]{3,})", nhead)
        if not m:
            m = re.search(r"\b(\d{1,2})\s+([a-z]{3,})\s+(\d{2,4})\b", nhead)
        if not m:
            return None
        day = int(m.group(1))
        mes_txt = m.group(2)[:3] if m.group(2)[:3] in MESES else m.group(2)
        month = MESES.get(mes_txt) or MESES.get(m.group(2))
        if not month:
            return None
        year = 2026
        if m.lastindex and m.lastindex >= 3:
            try:
                y = int(m.group(3))
                year = y if y > 100 else 2000 + y
            except (ValueError, IndexError):
                pass

    if not (1 <= day <= 31 and 1 <= month <= 12):
        return None

    hh, mm = None, None
    tm = TIME_RE.search(chunk[:80])
    if tm:
        try:
            hh = int(tm.group(1))
            mm = int(tm.group(2))
            ampm = (tm.group(3) or "").lower().replace(".", "").replace(" ", "")
            if ampm == "pm" and hh < 12:
                hh += 12
            if ampm == "am" and hh == 12:
                hh = 0
            if hh > 23:
                hh, mm = None, None
        except ValueError:
            hh, mm = None, None

    try:
        if hh is not None and mm is not None:
            return "%04d-%02d-%02d %02d:%02d" % (year, month, day, hh, mm)
        return "%04d-%02d-%02d" % (year, month, day)
    except ValueError:
        return None


def split_into_movements(text):
    if not text:
        return []
    text = str(text).strip()
    if not text:
        return []
    matches = list(DATE_LEAD_RE.finditer(text))
    if len(matches) < 2:
        return [{"fecha": parse_fecha_aproximada(text), "texto": text}]
    movimientos = []
    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        chunk = text[start:end].strip(" \n\t-*")
        if chunk:
            movimientos.append({"fecha": parse_fecha_aproximada(chunk), "texto": chunk})
    return movimientos


def process_workbook(path, file_country_map, log):
    fname = os.path.basename(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    file_default_country = file_country_map.get(fname, "")
    last_good_country = file_default_country

    companies = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row_idx, header_vals = find_header_row(ws)
        if header_row_idx is None:
            log.append(("HOJA SIN HEADER RECONOCIBLE", fname, sheet_name, "No se encontraron columnas Empresa/Nombre; hoja omitida"))
            continue

        extracted = extract_country_from_row1(ws)
        if extracted:
            sheet_country = extracted
            last_good_country = extracted
        else:
            sheet_country = last_good_country
        if not sheet_country:
            log.append(("HOJA SIN PAIS DETECTABLE", fname, sheet_name, "No se pudo determinar el pais/estado de esta hoja automaticamente; revisar y completar bp_region_comercial/country a mano"))

        roles = build_column_roles(header_vals)
        tz, ventana = extract_ref_timezone(ws)
        estado_prospeccion, motivo_descarte = classify_sheet(sheet_name)

        col_by_role = {}
        for i, role in enumerate(roles):
            if role and role not in col_by_role:
                col_by_role[role] = i + 1

        def val(row, role):
            c = col_by_role.get(role)
            if not c:
                return None
            return ws.cell(row=row, column=c).value

        current = None
        max_row = ws.max_row
        for r in range(header_row_idx + 1, max_row + 1):
            empresa_v = val(r, "empresa")
            nombre_v = val(r, "nombre_contacto")
            if empresa_v is None and nombre_v is None:
                continue
            if empresa_v is not None and str(empresa_v).strip():
                if current:
                    companies.append(current)
                prioridad_raw = val(r, "prioridad")
                prio_code, prio_country = split_priority_country(prioridad_raw)
                pais_row = prio_country or sheet_country
                current = {
                    "empresa": str(empresa_v).strip(),
                    "giro": val(r, "giro"),
                    "seguimiento_empresa": val(r, "seguimiento_empresa"),
                    "prioridad": prio_code,
                    "tel_empresa": val(r, "tel_empresa"),
                    "coment_tel_empresa": val(r, "coment_tel_empresa"),
                    "email_empresa": val(r, "email_empresa"),
                    "coment_email_empresa": val(r, "coment_email_empresa"),
                    "pagina_web": val(r, "pagina_web"),
                    "coment_historicos": val(r, "coment_historicos"),
                    "coment_roberto": val(r, "coment_roberto"),
                    "pais": pais_row,
                    "estado_prospeccion": estado_prospeccion,
                    "motivo_descarte": motivo_descarte,
                    "sheet_source": sheet_name,
                    "file_source": fname,
                    "tz": tz,
                    "ventana": ventana,
                    "contactos": [],
                }
            if nombre_v is not None and str(nombre_v).strip():
                if current is None:
                    log.append(("CONTACTO SIN EMPRESA PADRE", fname, sheet_name,
                                 "Fila %d: contacto '%s' sin empresa previa; omitido" % (r, nombre_v)))
                    continue
                cargo_v = val(r, "cargo")
                name_cargo_pairs = split_multi_names(nombre_v, cargo_v)
                for j, (one_name, one_cargo) in enumerate(name_cargo_pairs):
                    if not one_name:
                        continue
                    is_first = (j == 0)
                    current["contactos"].append({
                        "nombre": one_name,
                        "cargo": one_cargo,
                        "coment_persona": val(r, "coment_persona") if is_first else None,
                        "tel_movil": val(r, "tel_movil") if is_first else None,
                        "coment_llamada": val(r, "coment_llamada") if is_first else None,
                        "email_persona": val(r, "email_persona") if is_first else None,
                        "coment_email": val(r, "coment_email") if is_first else None,
                        "linkedin": val(r, "linkedin") if is_first else None,
                        "mensaje_lkd": val(r, "mensaje_lkd") if is_first else None,
                        "coment_lkd": val(r, "coment_lkd") if is_first else None,
                    })
                    if len(name_cargo_pairs) > 1 and not is_first:
                        log.append(("CONTACTO SIN DATOS PROPIOS (multi-nombre en una celda)", fname, sheet_name,
                                     "'%s' vino junto con otro(s) nombre(s) en la misma celda -- no se le asigno email/tel/comentarios de la fila por no poder saber a cual pertenecen; completar a mano si se necesita" % one_name))
        if current:
            companies.append(current)
    return companies


def resolve_country_and_region(pais_raw):
    key = norm(pais_raw)
    if key in STATE_TO_COUNTRY:
        return STATE_TO_COUNTRY[key], region_for(pais_raw), pais_raw
    return pais_raw, region_for(pais_raw), ""


def main():
    log = []
    all_companies = []
    files = sorted(glob.glob(os.path.join(SRC_DIR, "*.xlsx")))
    file_country_map = build_file_country_map(files)
    print("Procesando %d archivos..." % len(files))
    for f in files:
        comps = process_workbook(f, file_country_map, log)
        print("  %s: %d empresas" % (os.path.basename(f), len(comps)))
        all_companies.extend(comps)

    print("\nTotal empresas extraidas (antes de dedupe): %d" % len(all_companies))

    domain_registry = {}
    name_registry = {}
    empresas_rows = []
    contactos_rows = []
    deals_rows = []
    movimientos_rows = []
    idx = 0

    for comp in all_companies:
        domain = extract_domain(comp.get("email_empresa"),
                                 *[c.get("email_persona") for c in comp["contactos"]])
        if not domain and comp.get("pagina_web"):
            domain = normalize_website_domain(comp["pagina_web"])

        skip = False
        name_key = norm(comp["empresa"])

        if domain:
            existing_by_domain = domain_registry.get(domain, [])
            similar = next((n for n in existing_by_domain if similar_names(comp["empresa"], n)), None)
            if similar:
                log.append(("EMPRESA DUPLICADA (mismo dominio, nombre relacionado)", comp["file_source"], comp["sheet_source"],
                             "'%s' ya existia como '%s' con el mismo dominio (%s) -- omitida" % (comp["empresa"], similar, domain)))
                skip = True
            elif existing_by_domain:
                log.append(("DOMINIO COMPARTIDO ENTRE EMPRESAS DISTINTAS", comp["file_source"], comp["sheet_source"],
                             "'%s' comparte el dominio '%s' con %s (nombres no relacionados) -- se importan ambas, verificar" % (comp["empresa"], domain, existing_by_domain)))

        if not skip:
            existing_by_name = name_registry.get(name_key, [])
            conflict_free = next((e for e in existing_by_name if (not domain or not e["domain"] or domain == e["domain"])), None)
            if conflict_free:
                log.append(("EMPRESA DUPLICADA (mismo nombre)", comp["file_source"], comp["sheet_source"],
                             "'%s' ya existia (dominio previo: '%s', este: '%s') -- omitida" % (comp["empresa"], conflict_free["domain"], domain)))
                skip = True
            else:
                if existing_by_name:
                    log.append(("NOMBRE IDENTICO, DOMINIOS DISTINTOS", comp["file_source"], comp["sheet_source"],
                                 "'%s' tiene el mismo nombre que otra empresa ya vista pero con dominio distinto -- se importan ambas, verificar" % comp["empresa"]))
                name_registry.setdefault(name_key, []).append({"domain": domain, "name": comp["empresa"]})
                if domain:
                    domain_registry.setdefault(domain, []).append(comp["empresa"])

        if skip:
            continue

        idx += 1
        bp_id = "%s-%05d" % (OPERATOR_PREFIX, idx)

        if not domain:
            log.append(("SIN DOMINIO", comp["file_source"], comp["sheet_source"],
                         "'%s' (%s) sin email ni pagina web utilizable (o solo correo gratuito tipo gmail/hotmail) -- completar dominio a mano antes de importar" % (comp["empresa"], bp_id)))

        priority_norm = norm(comp.get("prioridad"))
        score = PRIORITY_SCORE.get(priority_norm, "")
        tipo_empresa = guess_tipo_empresa(comp.get("giro"))

        country_hubspot, region, state_tag = resolve_country_and_region(comp.get("pais"))
        giro_final = comp.get("giro") or ""
        if state_tag:
            giro_final = "[%s] %s" % (state_tag, giro_final)

        has_company_contact_data = bool(clean_phone(comp.get("tel_empresa"))) or bool(comp.get("email_empresa"))
        has_any_contact_with_data = any(
            (c.get("tel_movil") or c.get("email_persona")) for c in comp["contactos"]
        )
        estado_enriquecimiento = "completo" if (has_company_contact_data and has_any_contact_with_data) else (
            "parcial" if (has_company_contact_data or has_any_contact_with_data) else "pendiente")

        motivo_descarte = comp.get("motivo_descarte") or ""

        empresas_rows.append({
            "bp_id_unico": bp_id,
            "name": comp["empresa"],
            "bp_dominio_normalizado": domain,
            "domain": normalize_website_domain(comp.get("pagina_web")) if comp.get("pagina_web") else "",
            "country": country_hubspot or "",
            "bp_region_comercial": region,
            "bp_zona": OPERATOR_ZONE,
            "bp_tipo_empresa": tipo_empresa,
            "bp_estado_prospeccion": "nueva" if comp["estado_prospeccion"] == "activo" else "no_participa",
            "bp_estado_enriquecimiento": estado_enriquecimiento,
            "bp_fuente_enriquecimiento": "manual",
            "bp_clasificacion_claude": "",
            "bp_score_bepharma": score,
            "bp_estado_alerta": "",
            "bp_zona_horaria_pais": comp.get("tz") or "",
            "bp_ventana_llamada_mx": comp.get("ventana") or "",
            "phone": clean_phone(comp.get("tel_empresa")),
            "industry": "",
            "numberofemployees": "",
            "bp_motivo_descarte": motivo_descarte,
            "description": giro_final,
        })

        for c in comp["contactos"]:
            if not (c.get("nombre") or "").strip():
                continue
            firstname, lastname = split_name(c["nombre"])
            email = (c.get("email_persona") or "").strip() if c.get("email_persona") else ""
            invitacion = "enviada" if (c.get("coment_email") or comp.get("coment_email_empresa")) else "no_enviada"
            contactos_rows.append({
                "firstname": firstname,
                "lastname": lastname,
                "email": email,
                "phone": clean_phone(c.get("tel_movil")),
                "jobtitle": c.get("cargo") or "",
                "company": comp["empresa"],
                "associatedcompanydomain": domain,
                "bp_cargo_decisor": guess_cargo_decisor(c.get("cargo")),
                "bp_email_verificado": "en_verificacion" if email else "",
                "bp_telefono_verificado": "en_verificacion" if clean_phone(c.get("tel_movil")) else "",
                "bp_fuente_contacto": "manual",
                "bp_estado_invitacion": invitacion,
                "linkedin_url": c.get("linkedin") or "",
            })
            if not email:
                log.append(("CONTACTO SIN EMAIL", comp["file_source"], comp["sheet_source"],
                            "'%s' en '%s' sin email -- HubSpot puede rechazar el contacto si no tiene ningun identificador" % (c["nombre"], comp["empresa"])))

            for canal, texto in (
                ("Comentario contacto", c.get("coment_persona")),
                ("Llamada contacto", c.get("coment_llamada")),
                ("Email contacto", c.get("coment_email")),
                ("LinkedIn mensaje", c.get("mensaje_lkd")),
                ("LinkedIn comentario", c.get("coment_lkd")),
            ):
                for mov in split_into_movements(texto):
                    movimientos_rows.append({
                        "dealname": "%s - %s" % (comp["empresa"], EVENT_CODE),
                        "associatedcompanydomain": domain,
                        "empresa": comp["empresa"],
                        "contacto_relacionado": c["nombre"],
                        "canal": canal,
                        "hs_timestamp": mov["fecha"] or "",
                        "hs_note_body": "[%s | %s] %s" % (canal, c["nombre"], mov["texto"]),
                        "hubspot_owner_id": OPERATOR_EMAIL,
                    })
                    if not mov["fecha"]:
                        log.append(("MOVIMIENTO SIN FECHA RECONOCIDA", comp["file_source"], comp["sheet_source"],
                                     "Contacto '%s' en '%s' (%s): no se pudo detectar fecha -- completar a mano" % (c["nombre"], comp["empresa"], canal)))

        dealstage_label = "Nueva empresa" if comp["estado_prospeccion"] == "activo" else "No participa"
        deals_rows.append({
            "dealname": "%s - %s" % (comp["empresa"], EVENT_CODE),
            "associatedcompanydomain": domain,
            "pipeline": PIPELINE_ID,
            "dealstage": dealstage_label,
            "hubspot_owner_id": OPERATOR_EMAIL,
            "bp_evento_codigo": EVENT_CODE,
            "bp_evento_nombre": "",
            "bp_evento_estado": "",
            "bp_zona": OPERATOR_ZONE,
            "bp_evento_fecha_inicio": "",
            "bp_evento_fecha_cierre": "",
            "bp_evento_meta_contactar": "",
            "bp_evento_meta_confirmar": "",
            "bp_evento_paises": "",
            "closedate": "",
        })

        for canal, texto in (
            ("Seguimiento general", comp.get("seguimiento_empresa")),
            ("Llamada empresa", comp.get("coment_tel_empresa")),
            ("Email empresa", comp.get("coment_email_empresa")),
            ("Historico", comp.get("coment_historicos")),
            ("Notas Roberto", comp.get("coment_roberto")),
        ):
            for mov in split_into_movements(texto):
                movimientos_rows.append({
                    "dealname": "%s - %s" % (comp["empresa"], EVENT_CODE),
                    "associatedcompanydomain": domain,
                    "empresa": comp["empresa"],
                    "contacto_relacionado": "",
                    "canal": canal,
                    "hs_timestamp": mov["fecha"] or "",
                    "hs_note_body": "[%s] %s" % (canal, mov["texto"]),
                    "hubspot_owner_id": OPERATOR_EMAIL,
                })
                if not mov["fecha"]:
                    log.append(("MOVIMIENTO SIN FECHA RECONOCIDA", comp["file_source"], comp["sheet_source"],
                                 "Empresa '%s' (%s): no se pudo detectar fecha -- completar a mano" % (comp["empresa"], canal)))

    print("\nEmpresas finales (post-dedupe): %d" % len(empresas_rows))
    print("Contactos: %d" % len(contactos_rows))
    print("Deals: %d" % len(deals_rows))
    print("Alertas en log: %d" % len(log))
    print("Movimientos: %d" % len(movimientos_rows))

    write_output(empresas_rows, contactos_rows, deals_rows, movimientos_rows, log)


def write_output(empresas_rows, contactos_rows, deals_rows, movimientos_rows, log):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def add_sheet(name, rows, columns):
        ws = wb.create_sheet(name)
        ws.append(columns)
        for row in rows:
            ws.append([row.get(col, "") for col in columns])
        for i, col in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(14, min(40, len(col) + 4))
        ws.freeze_panes = "A2"

    empresas_cols = ["bp_id_unico", "name", "bp_dominio_normalizado", "domain", "country",
                     "bp_region_comercial", "bp_zona", "bp_tipo_empresa", "bp_estado_prospeccion",
                     "bp_estado_enriquecimiento", "bp_fuente_enriquecimiento", "bp_clasificacion_claude",
                     "bp_score_bepharma", "bp_estado_alerta", "bp_zona_horaria_pais",
                     "bp_ventana_llamada_mx", "phone", "industry", "numberofemployees",
                     "bp_motivo_descarte", "description"]
    contactos_cols = ["firstname", "lastname", "email", "phone", "jobtitle", "company",
                      "associatedcompanydomain", "bp_cargo_decisor", "bp_email_verificado",
                      "bp_telefono_verificado", "bp_fuente_contacto", "bp_estado_invitacion",
                      "linkedin_url"]
    deals_cols = ["dealname", "associatedcompanydomain", "pipeline", "dealstage",
                  "hubspot_owner_id", "bp_evento_codigo", "bp_evento_nombre", "bp_evento_estado",
                  "bp_zona", "bp_evento_fecha_inicio", "bp_evento_fecha_cierre",
                  "bp_evento_meta_contactar", "bp_evento_meta_confirmar", "bp_evento_paises", "closedate"]
    movimientos_cols = ["dealname", "associatedcompanydomain", "empresa", "contacto_relacionado",
                        "canal", "hs_timestamp", "hs_note_body", "hubspot_owner_id"]

    add_sheet("EMPRESAS", empresas_rows, empresas_cols)
    add_sheet("CONTACTOS", contactos_rows, contactos_cols)
    add_sheet("DEALS", deals_rows, deals_cols)
    add_sheet("MOVIMIENTOS_DEAL", movimientos_rows, movimientos_cols)

    ws_log = wb.create_sheet("REVISAR ANTES DE IMPORTAR")
    ws_log.append(["Tipo de alerta", "Archivo", "Hoja", "Detalle"])
    for row in log:
        ws_log.append(list(row))
    for i, w in enumerate([35, 30, 25, 90], start=1):
        ws_log.column_dimensions[get_column_letter(i)].width = w
    ws_log.freeze_panes = "A2"

    wb.save(OUT_XLSX)
    print("\nArchivo guardado: %s" % OUT_XLSX)


if __name__ == "__main__":
    main()
